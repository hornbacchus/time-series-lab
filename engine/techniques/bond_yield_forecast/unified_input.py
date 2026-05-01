"""Unified input-workbook architecture (Step 5).

A single ``bvar_inputs.xlsx`` workbook contains all history (macro +
yields sheets) plus up to five projection scenarios
(``projections_baseline`` plus ``projections_scenario_1`` ..
``projections_scenario_4``) in one file. The desk's quarterly cycle
populates this workbook once per quarter; the canonical CLI command
reads it, re-estimates the BVAR-SV from the workbook's history, and
runs conditional forecasts for one named scenario or all populated
scenarios.

This module provides:

- ``read_unified_workbook(path, scenario, config)`` — read history +
  one named scenario, returning validated DataFrames + an
  ``EconomistProjections`` object.
- ``write_unified_template(output_path, horizon, copy_history_from,
  config)`` — generate the 8-sheet template, optionally copying
  history from an existing workbook (migration helper from the old
  two-file workflow).
- ``list_available_scenarios(path, config)`` — return the list of
  scenario sheets that contain at least one populated data row.

No new model code, no new sampler logic. The Step 4
``ConditionalForecaster`` and Step 2 ``BVARSV`` are reused unchanged.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from .conditioning import EconomistProjections
from .data import InputValidationError

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Sheet-name resolution
# ---------------------------------------------------------------------------


_DEFAULT_SHEET_NAMES = {
    "macro_history": "macro",
    "yields_history": "yields",
    "projections_baseline": "projections_baseline",
    "projections_scenario_1": "projections_scenario_1",
    "projections_scenario_2": "projections_scenario_2",
    "projections_scenario_3": "projections_scenario_3",
    "projections_scenario_4": "projections_scenario_4",
    "readme": "README",
}

_SCENARIO_KEYS = (
    "baseline",
    "scenario_1",
    "scenario_2",
    "scenario_3",
    "scenario_4",
)


def _resolve_sheet_names(config: dict) -> dict[str, str]:
    """Resolve the unified-workbook sheet names from config with defaults."""
    cfg = (config or {}).get("unified_input", {}).get("sheet_names", {})
    return {**_DEFAULT_SHEET_NAMES, **cfg}


def _scenario_sheet_name(scenario: str, sheet_names: dict[str, str]) -> str:
    """Map a scenario key (e.g. 'baseline') to its workbook sheet name."""
    if scenario == "baseline":
        return sheet_names["projections_baseline"]
    if scenario in ("scenario_1", "scenario_2", "scenario_3", "scenario_4"):
        return sheet_names[f"projections_{scenario}"]
    raise ValueError(
        f"Unknown scenario {scenario!r}; expected one of "
        f"{('baseline',) + tuple(f'scenario_{i}' for i in range(1, 5))}"
    )


# ---------------------------------------------------------------------------
# Reader: read_unified_workbook
# ---------------------------------------------------------------------------


def _is_data_row_populated(row: pd.Series, expected_macro_cols: list[str]) -> bool:
    """Return True if the row has any non-blank, non-NaN macro cell."""
    for c in expected_macro_cols:
        if c not in row.index:
            continue
        val = row[c]
        if val is None:
            continue
        if isinstance(val, float) and np.isnan(val):
            continue
        if isinstance(val, str) and val.strip() == "":
            continue
        return True
    return False


def list_available_scenarios(
    path: str | Path, config: dict
) -> list[str]:
    """Return scenario keys whose sheets contain at least one populated data row.

    Always includes 'baseline' first (if populated). Numbered scenarios
    follow in order. A sheet is "populated" iff at least one data row
    has any non-blank macro cell.
    """
    path = Path(path)
    sheet_names = _resolve_sheet_names(config)
    macro_specs = config["data"]["macro_variables"]
    expected_macro_cols = [spec["column"] for spec in macro_specs.values()]

    xl = pd.ExcelFile(path, engine="openpyxl")
    available: list[str] = []
    for key in _SCENARIO_KEYS:
        sheet = _scenario_sheet_name(key, sheet_names)
        if sheet not in xl.sheet_names:
            continue
        try:
            df = pd.read_excel(xl, sheet_name=sheet)
        except Exception as exc:  # pragma: no cover — pandas decides
            logger.warning("Failed to read sheet '%s': %s", sheet, exc)
            continue
        if df.shape[0] == 0:
            continue
        # Require at least one row populated in any of the expected macro columns.
        any_populated = False
        for _, row in df.iterrows():
            if _is_data_row_populated(row, expected_macro_cols):
                any_populated = True
                break
        if any_populated:
            available.append(key)
    return available


def read_unified_workbook(
    path: str | Path,
    scenario: str,
    config: dict,
) -> dict[str, Any]:
    """Read history + one named projection scenario from the unified workbook.

    Returns a dict::

        {
          "macro_history":  pd.DataFrame (PeriodIndex, config-key columns),
          "yields_history": pd.DataFrame (PeriodIndex, config-key columns),
          "projections":    EconomistProjections,
          "raw":            {"macro_raw": df, "yields_raw": df},  # for build_panel
          "metadata":       {
              "workbook_path": str,
              "scenario": str,
              "macro_history_quarters": [str, str],
              "projection_quarters":    [str, str],
              "available_scenarios":    list[str],
          },
        }

    Raises ``InputValidationError`` on:
      - missing required sheets,
      - empty (header-only) chosen scenario sheet,
      - history alignment / format issues (delegated to data.validate_input),
      - projection alignment / format issues (delegated to
        EconomistProjections.validate).
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(
            f"Unified input workbook not found at {path}. "
            f"Generate via: python -m cli.run_forecast "
            f"--generate-input-template --output {path}"
        )

    sheet_names = _resolve_sheet_names(config)
    scenario_sheet = _scenario_sheet_name(scenario, sheet_names)

    xl = pd.ExcelFile(path, engine="openpyxl")
    available_xl = set(xl.sheet_names)

    required = [sheet_names["macro_history"], sheet_names["yields_history"]]
    missing = [s for s in required + [scenario_sheet] if s not in available_xl]
    if missing:
        raise InputValidationError(
            "missing_sheet",
            f"Unified workbook missing required sheet(s): {missing}. "
            f"Available: {sorted(available_xl)}.",
        )

    # Read history sheets (raw — Excel headers, Quarter index as strings).
    macro_raw = pd.read_excel(xl, sheet_name=sheet_names["macro_history"])
    yields_raw = pd.read_excel(xl, sheet_name=sheet_names["yields_history"])

    for name, df in [("macro", macro_raw), ("yields", yields_raw)]:
        if df.columns.size == 0 or df.columns[0] != "Quarter":
            first = df.columns[0] if df.columns.size > 0 else "(empty)"
            raise InputValidationError(
                "missing_quarter_column",
                f"First column on '{name}' sheet must be 'Quarter'; got '{first}'",
                location=f"sheet='{name}'",
            )

    macro_raw = macro_raw.set_index("Quarter")
    yields_raw = yields_raw.set_index("Quarter")
    raw = {"macro_raw": macro_raw, "yields_raw": yields_raw}

    # Validate history through the existing data.validate_input pipeline.
    from .data import validate_input
    clean = validate_input(raw, config)
    macro_history = clean["macro"]
    yields_history = clean["yields"]

    # Read scenario sheet.
    scenario_raw = pd.read_excel(xl, sheet_name=scenario_sheet)
    macro_specs = config["data"]["macro_variables"]
    expected_macro_cols = [spec["column"] for spec in macro_specs.values()]
    populated_rows = [
        row for _, row in scenario_raw.iterrows()
        if _is_data_row_populated(row, expected_macro_cols)
    ]
    if not populated_rows:
        raise InputValidationError(
            "empty_scenario",
            f"Scenario {scenario!r} is empty in the input workbook. "
            f"Fill in the projections sheet or choose a different scenario.",
            location=f"sheet='{scenario_sheet}'",
        )

    # Strip blank trailing rows (where Quarter is NaN) before passing to
    # EconomistProjections.from_dataframe.
    scenario_clean = scenario_raw[scenario_raw["Quarter"].notna()].reset_index(drop=True)
    projections = EconomistProjections.from_dataframe(
        scenario_clean, config, sheet_name=scenario_sheet
    )

    # Cross-validate alignment: projections begin at history.last_period + 1.
    projections.validate(macro_history, n_lags=int(config["model"]["lags"]))

    available = list_available_scenarios(path, config)

    metadata = {
        "workbook_path": str(path),
        "scenario": scenario,
        "macro_history_quarters": [
            str(macro_history.index[0]), str(macro_history.index[-1])
        ],
        "projection_quarters": [
            str(projections.projections.index[0]),
            str(projections.projections.index[-1]),
        ],
        "available_scenarios": available,
    }

    return {
        "macro_history": macro_history,
        "yields_history": yields_history,
        "projections": projections,
        "raw": raw,
        "metadata": metadata,
    }


# ---------------------------------------------------------------------------
# Writer: write_unified_template
# ---------------------------------------------------------------------------


_README_LINES: list[tuple[str, bool]] = [
    ("BVAR Yield Forecaster — Unified Input Workbook", True),
    ("", False),
    (
        "This workbook is the canonical input artifact for the BVAR yield "
        "forecaster (Step 5 architecture). It contains historical macro and "
        "yield data plus up to five forward-looking projection scenarios. The "
        "pipeline reads everything from this single file.",
        False,
    ),
    ("", False),
    ("REQUIRED SHEETS", True),
    ("- macro                    : historical macro variables (one column per variable).", False),
    ("- yields                   : historical Treasury yields (one column per maturity).", False),
    ("- projections_baseline     : forward macro projections — REQUIRED.", False),
    ("- projections_scenario_1   : optional alternate scenario.", False),
    ("- projections_scenario_2   : optional alternate scenario.", False),
    ("- projections_scenario_3   : optional alternate scenario.", False),
    ("- projections_scenario_4   : optional alternate scenario.", False),
    ("- README                   : this sheet — ignored by the pipeline.", False),
    ("", False),
    ("QUARTER FORMAT (all sheets)", True),
    ("- 'Quarter' is always the first column.", False),
    ("- Values match the regex ^\\d{4}-Q[1-4]$ (e.g. 1990-Q1, 2026-Q1).", False),
    ("- All values in PERCENT, never decimal form.", False),
    ("", False),
    ("HISTORY SHEETS — macro and yields", True),
    ("- Same Quarter index across both sheets.", False),
    ("- Monotonically increasing, no duplicates, no interior gaps.", False),
    ("- Required macro columns:", False),
    ("    Real GDP Growth (Q/Q SAAR)", False),
    ("    Headline CPI Inflation (Q/Q annualized)", False),
    ("    Fed Funds Rate (quarterly average)", False),
    ("- Required yields columns: 3M, 6M, 1Y, 2Y, 3Y, 5Y, 7Y, 10Y, 20Y, 30Y.", False),
    ("", False),
    ("PROJECTION SHEETS — projections_baseline + scenario_1..4", True),
    ("- Quarters begin IMMEDIATELY after the last quarter in 'macro' / 'yields'.", False),
    ("  No gaps, no overlaps with history.", False),
    ("- Same three macro columns as the macro sheet.", False),
    ("- All values in PERCENT.", False),
    ("- Values must be in [-50, 50] (rejects unit-scale errors).", False),
    ("- A sheet is 'populated' if it has at least one data row with a non-blank", False),
    ("  macro cell. Header-only sheets are treated as 'empty' and skipped by", False),
    ("  --scenario all; an explicit --scenario X on an empty sheet errors out.", False),
    ("", False),
    ("USAGE", True),
    ("Single-scenario run:", False),
    ("    python -m cli.run_forecast --forecast \\", False),
    ("        --input data/raw/bvar_inputs.xlsx \\", False),
    ("        --scenario baseline", False),
    ("", False),
    ("All-populated-scenarios run (single estimation, multiple forecasts):", False),
    ("    python -m cli.run_forecast --forecast \\", False),
    ("        --input data/raw/bvar_inputs.xlsx \\", False),
    ("        --scenario all", False),
    ("", False),
    ("List which scenarios are populated, without running:", False),
    ("    python -m cli.run_forecast --list-scenarios \\", False),
    ("        --input data/raw/bvar_inputs.xlsx", False),
]


def _write_readme_sheet(ws) -> None:
    from openpyxl.styles import Alignment, Font
    bold = Font(bold=True, size=12)
    wrap = Alignment(wrap_text=True, vertical="top")
    for row_idx, (text, is_bold) in enumerate(_README_LINES, start=1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.alignment = wrap
        if is_bold:
            cell.font = bold
    ws.column_dimensions["A"].width = 90


def _quarter_label(period: pd.Period) -> str:
    """Format a Period as YYYY-Q# matching the validation regex."""
    return f"{period.year}-Q{period.quarter}"


def _last_period_from_macro(macro_df: pd.DataFrame) -> pd.Period | None:
    """Best-effort extraction of the most recent quarter from a macro DataFrame.

    Accepts either a string-indexed (raw) DataFrame or a PeriodIndex one.
    Returns None if the index is empty.
    """
    if macro_df.shape[0] == 0:
        return None
    last = macro_df.index[-1]
    if isinstance(last, pd.Period):
        return last
    if isinstance(last, str):
        try:
            return pd.Period(last, freq="Q-DEC")
        except Exception:  # pragma: no cover
            return None
    return None


def write_unified_template(
    output_path: str | Path,
    horizon: int,
    copy_history_from: str | Path | None = None,
    config: dict | None = None,
) -> Path:
    """Generate the 8-sheet unified template workbook.

    If ``copy_history_from`` is provided, the macro and yields sheets are
    copied verbatim from that workbook (raw — original Excel headers and
    string Quarter labels). The five projection sheets are pre-populated
    with Quarter labels for the next ``horizon`` quarters relative to the
    most recent quarter present in the history.

    If no source workbook is given, the macro and yields sheets are
    header-only; quarter labels for the projection sheets start at the
    current quarter (today's date converted to a Q-DEC Period).
    """
    from openpyxl import Workbook, load_workbook

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if config is None:
        # Sensible defaults: load the project default config if accessible.
        from .data import load_config as _load_config
        try:
            config = _load_config(Path("config/default.yaml"))
        except Exception:  # pragma: no cover — fall through to minimal defaults
            config = {"data": {"macro_variables": {}, "yield_variables": {}}}

    macro_specs = config["data"]["macro_variables"]
    yield_specs = config["data"]["yield_variables"]
    macro_columns = [spec["column"] for spec in macro_specs.values()]
    yield_columns = [spec["column"] for spec in yield_specs.values()]
    sheet_names = _resolve_sheet_names(config)

    wb = Workbook()
    # Remove the default initial sheet so we can add named sheets in order.
    default_ws = wb.active
    wb.remove(default_ws)

    # macro / yields history sheets
    macro_ws = wb.create_sheet(sheet_names["macro_history"])
    yields_ws = wb.create_sheet(sheet_names["yields_history"])

    last_period: pd.Period | None = None
    if copy_history_from is not None:
        src_path = Path(copy_history_from)
        if not src_path.exists():
            raise FileNotFoundError(
                f"copy_history_from source not found: {src_path}"
            )
        # Read source macro and yields. Try the unified sheet names first;
        # if not present, fall back to the legacy ones from
        # config["data"]["macro_sheet"] / ["yield_sheet"].
        legacy_macro_sheet = config["data"].get("macro_sheet", "macro")
        legacy_yields_sheet = config["data"].get("yield_sheet", "yields")
        src_xl = pd.ExcelFile(src_path, engine="openpyxl")
        macro_src_sheet = (
            sheet_names["macro_history"]
            if sheet_names["macro_history"] in src_xl.sheet_names
            else legacy_macro_sheet
        )
        yields_src_sheet = (
            sheet_names["yields_history"]
            if sheet_names["yields_history"] in src_xl.sheet_names
            else legacy_yields_sheet
        )
        if macro_src_sheet not in src_xl.sheet_names:
            raise InputValidationError(
                "missing_sheet",
                f"copy-history source missing macro sheet "
                f"('{macro_src_sheet}'); available: {src_xl.sheet_names}",
            )
        if yields_src_sheet not in src_xl.sheet_names:
            raise InputValidationError(
                "missing_sheet",
                f"copy-history source missing yields sheet "
                f"('{yields_src_sheet}'); available: {src_xl.sheet_names}",
            )
        src_macro = pd.read_excel(src_xl, sheet_name=macro_src_sheet)
        src_yields = pd.read_excel(src_xl, sheet_name=yields_src_sheet)
        # Write rows (header + data) to the new workbook.
        macro_ws.append(list(src_macro.columns))
        for _, row in src_macro.iterrows():
            macro_ws.append([row[col] for col in src_macro.columns])
        yields_ws.append(list(src_yields.columns))
        for _, row in src_yields.iterrows():
            yields_ws.append([row[col] for col in src_yields.columns])
        # Determine last period from the source macro sheet's Quarter column.
        if "Quarter" in src_macro.columns and src_macro.shape[0] > 0:
            try:
                last_q_str = str(src_macro["Quarter"].iloc[-1])
                last_period = pd.Period(last_q_str, freq="Q-DEC")
            except Exception:
                last_period = None
    else:
        # Header-only history sheets.
        macro_ws.append(["Quarter"] + macro_columns)
        yields_ws.append(["Quarter"] + yield_columns)

    if last_period is None:
        last_period = pd.Period(pd.Timestamp.today(), freq="Q-DEC")

    # Five projection sheets, all with the same shape.
    projection_quarters = pd.period_range(
        last_period + 1, periods=horizon, freq="Q-DEC"
    )
    quarter_labels = [_quarter_label(q) for q in projection_quarters]
    projection_sheet_keys = [
        "projections_baseline",
        "projections_scenario_1",
        "projections_scenario_2",
        "projections_scenario_3",
        "projections_scenario_4",
    ]
    for key in projection_sheet_keys:
        ws = wb.create_sheet(sheet_names[key])
        ws.append(["Quarter"] + macro_columns)
        for label in quarter_labels:
            ws.append([label] + ["" for _ in macro_columns])

    # README sheet last.
    readme_ws = wb.create_sheet(sheet_names["readme"])
    _write_readme_sheet(readme_ws)

    wb.save(output_path)
    logger.info(
        "Wrote unified-input template to %s (history sheets: %s, "
        "projection scenarios: %s, horizon=%d, last_period=%s)",
        output_path,
        "copied from " + str(copy_history_from) if copy_history_from else "header-only",
        len(projection_sheet_keys),
        horizon,
        last_period,
    )
    return output_path
