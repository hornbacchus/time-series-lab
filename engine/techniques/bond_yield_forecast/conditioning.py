"""BGL-2015 / Waggoner-Zha-1999 conditional forecasting for BVAR-SV.

Algorithm summary
-----------------
For each posterior draw d in BVARSVResults, for each path p:
  - Forward-simulate the joint (target, macro) path one horizon at a time.
  - At each horizon t, condition on the economist projection of the
    macro variables:
      * Strict match (default): macro_t = projection_t exactly; sample
        target_t from the conditional Gaussian given the projection.
      * Soft match: treat projection_t = macro_t + eta, eta ~ N(0, R);
        sample (target_t, macro_t) jointly from the posterior given
        the projection.
  - Reassemble y_t, advance the state, evolve h_t -> h_{t+1}.

This per-time-t conditional Gaussian approach is the natural BVAR-SV
generalization of Waggoner-Zha-1999. Under SV the per-horizon
covariance is path-dependent through h_t, so the constant-vol stacked
across-horizon covariance does not apply. Each draw and each path
contributes one conditional path; the ensemble of paths IS the
conditional posterior predictive distribution.

The volatility-evolution timing matches
``BVARSVResults.posterior_predictive_unconditional`` exactly: h_T
applies at horizon t=0 (i.e., the first new period), and h is advanced
*after* each step. This timing is essential so that a conditional
run with the projection set to the unconditional median agrees with
the unconditional posterior predictive in the limit of large
projection uncertainty.

Citations
---------
    Banbura, M., Giannone, D., Lenza, M. (2015). "Conditional
        forecasts and scenario analysis with vector autoregressions
        for large cross-sections." J. Applied Econometrics
        30(7):739-756.
    Waggoner, D. and Zha, T. (1999). "Conditional forecasts in
        dynamic multivariate models." RES 81(4):639-651.
    Doan, T., Litterman, R., Sims, C. (1984). "Forecasting and
        conditional projection using realistic prior distributions."
        Econometric Reviews 3:1-100.
"""

from __future__ import annotations

import json
import logging
import re
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from ._conditional_inner import conditional_forecast_inner_loop
from .data import InputValidationError, load_input_workbook
from .estimation import BVARSVResults, _to_jsonable

logger = logging.getLogger(__name__)

_QUARTER_RE = re.compile(r"^\d{4}-Q[1-4]$")

# Sanity bounds reused for projection validation. Wider than the
# input-workbook macro bounds so the desk can project unusual scenarios
# (e.g. a deep recession or strong tightening cycle) without tripping.
_PROJECTION_BOUNDS = (-50.0, 50.0)

# S0.3 Refinement 1 — tiered guard thresholds for projection_uncertainty.
# Below the hard threshold, the Schur complement becomes near-singular
# regardless of covariance shape; raise ValueError. Between hard and soft,
# stability is shape-dependent; emit BVARWarning so the user can decide.
_HARD_MIN_PROJECTION_UNCERTAINTY = 1e-4
_SOFT_MIN_PROJECTION_UNCERTAINTY = 1e-2


# ---------------------------------------------------------------------------
# Small helpers (mirror data.py without forcing a public API change there)
# ---------------------------------------------------------------------------


def _check_quarter_format(index: pd.Index, sheet: str) -> None:
    for q in index:
        if not isinstance(q, str) or not _QUARTER_RE.match(q):
            raise InputValidationError(
                "bad_quarter_format",
                f"Quarter value {q!r} does not match YYYY-Q# format on sheet '{sheet}'",
                location=f"sheet='{sheet}', Quarter={q!r}",
            )


def _to_period_index(string_index: pd.Index, sheet: str) -> pd.PeriodIndex:
    try:
        return pd.PeriodIndex(
            [pd.Period(q, freq="Q-DEC") for q in string_index], freq="Q-DEC"
        )
    except Exception as exc:
        raise InputValidationError(
            "bad_quarter_format",
            f"Failed to parse Quarter values on sheet '{sheet}': {exc}",
            location=f"sheet='{sheet}'",
        ) from exc


def _symmetrize(M: np.ndarray) -> np.ndarray:
    """Symmetrize and floor tiny negative eigenvalues from numerical noise."""
    M = 0.5 * (M + M.T)
    # Add a small ridge to guard rng.multivariate_normal's Cholesky path on
    # near-singular conditional covariances.
    return M + 1e-12 * np.eye(M.shape[0])


# ---------------------------------------------------------------------------
# EconomistProjections
# ---------------------------------------------------------------------------


class EconomistProjections:
    """Container for desk-supplied future macro paths.

    The DataFrame uses a quarterly ``PeriodIndex`` and column names matching
    the canonical config keys (``real_gdp_growth``, ``headline_cpi``,
    ``fed_funds_rate``), in the order specified by ``macro_variables``.
    """

    def __init__(
        self,
        projections: pd.DataFrame,
        macro_variables: list[str],
    ):
        if not isinstance(projections.index, pd.PeriodIndex):
            raise InputValidationError(
                "bad_projection_index",
                f"projections must have PeriodIndex; got {type(projections.index).__name__}",
            )
        missing = [v for v in macro_variables if v not in projections.columns]
        if missing:
            raise InputValidationError(
                "missing_column",
                f"Projections missing required column(s): {missing}",
            )
        # Reorder to the canonical macro_variables order.
        self.projections = projections[macro_variables].copy()
        self.macro_variables = list(macro_variables)

    @property
    def horizon(self) -> int:
        return int(self.projections.shape[0])

    def to_array(self, macro_variables: list[str]) -> np.ndarray:
        """Return a (horizon, n_macro) array in the requested column order."""
        missing = [v for v in macro_variables if v not in self.projections.columns]
        if missing:
            raise KeyError(
                f"requested macro_variables {missing} not in projections "
                f"(have {list(self.projections.columns)})"
            )
        return self.projections[macro_variables].to_numpy(dtype=float)

    def validate(self, panel: pd.DataFrame, n_lags: int) -> None:
        """Validate alignment with the in-sample panel.

        Raises ``InputValidationError`` on:
          - first projection period != panel.index[-1] + 1 (gap or overlap)
          - projection PeriodIndex not contiguous quarterly
          - any NaN in the projection grid
          - any value outside ``_PROJECTION_BOUNDS`` (rejects unit-scale errors)
        """
        if not isinstance(panel.index, pd.PeriodIndex):
            raise InputValidationError(
                "bad_panel_index",
                f"panel must have PeriodIndex; got {type(panel.index).__name__}",
            )
        if self.horizon == 0:
            raise InputValidationError(
                "empty_projections",
                "projections workbook contains zero rows",
            )

        last = panel.index[-1]
        first_proj = self.projections.index[0]
        expected_first = last + 1
        if first_proj != expected_first:
            raise InputValidationError(
                "projection_alignment",
                f"first projection period must equal panel.last + 1 = {expected_first}; "
                f"got {first_proj} (panel.last = {last}). Gap or overlap detected.",
                location=f"first_projection_period={first_proj}",
            )

        # Contiguous, quarterly.
        idx = self.projections.index
        for i in range(1, len(idx)):
            if idx[i] != idx[i - 1] + 1:
                raise InputValidationError(
                    "projection_quarter_gap",
                    f"projection quarters must be contiguous; gap at {idx[i - 1]} -> {idx[i]}",
                    location=f"projections row {i}",
                )

        # NaN check.
        if self.projections.isna().any().any():
            bad_cells: list[str] = []
            for col in self.projections.columns:
                for q in self.projections.index[self.projections[col].isna()]:
                    bad_cells.append(f"{col}@{q}")
                    if len(bad_cells) >= 3:
                        break
                if len(bad_cells) >= 3:
                    break
            raise InputValidationError(
                "projection_nan",
                f"projections contain NaN: {bad_cells}",
            )

        # Sanity bounds.
        lo, hi = _PROJECTION_BOUNDS
        out_of_range = (self.projections < lo) | (self.projections > hi)
        if out_of_range.any().any():
            for col in self.projections.columns:
                bad = self.projections[out_of_range[col]][col]
                if not bad.empty:
                    q = bad.index[0]
                    val = float(bad.iloc[0])
                    raise InputValidationError(
                        "projection_out_of_range",
                        f"projection {col}={val:.4g} at {q} outside [{lo}, {hi}]; "
                        "did you supply decimal form instead of percent?",
                        location=f"column='{col}', quarter={q}",
                    )

        # Lag count is informational only here — the conditional algorithm
        # uses it via the panel, but projection validation does not depend
        # on it. Include the parameter in the signature so callers can
        # express intent uniformly with other validators.
        del n_lags

    @classmethod
    def from_dataframe(
        cls,
        df: pd.DataFrame,
        config: dict,
        sheet_name: str = "projections",
    ) -> "EconomistProjections":
        """Build an EconomistProjections from an already-loaded DataFrame.

        Performs the same column-mapping, format, and numeric checks as
        ``from_workbook``. Use this from any reader that already has the
        DataFrame in hand (e.g., the unified-workbook reader). The
        DataFrame is expected to be raw — index = Quarter strings, columns
        = the human-readable Excel headers from
        ``config["data"]["macro_variables"][*]["column"]``.
        """
        if df.columns.size == 0 or df.columns[0] != "Quarter":
            first = df.columns[0] if df.columns.size > 0 else "(empty)"
            raise InputValidationError(
                "missing_quarter_column",
                f"First column on sheet '{sheet_name}' must be 'Quarter'; got '{first}'",
                location=f"sheet='{sheet_name}'",
            )

        df = df.set_index("Quarter")

        macro_specs = config["data"]["macro_variables"]
        canonical_keys = list(macro_specs.keys())
        col_to_key = {spec["column"]: key for key, spec in macro_specs.items()}

        missing_cols = [col for col in col_to_key if col not in df.columns]
        if missing_cols:
            raise InputValidationError(
                "missing_column",
                f"Projections sheet missing required column(s): {missing_cols}",
                location=f"sheet='{sheet_name}'",
            )

        extra = [c for c in df.columns if c not in col_to_key]
        if extra:
            logger.warning(
                "Extra columns in projections sheet ignored: %s", extra
            )

        df = df[list(col_to_key.keys())]
        _check_quarter_format(df.index, sheet_name)
        df.index = _to_period_index(df.index, sheet_name)

        for col in df.columns:
            converted = pd.to_numeric(df[col], errors="coerce")
            new_nan_mask = df[col].notna() & converted.isna()
            if new_nan_mask.any():
                first_q = df.index[new_nan_mask][0]
                bad_val = df.loc[first_q, col]
                raise InputValidationError(
                    "non_numeric",
                    f"Non-numeric value {bad_val!r} in projections column '{col}'",
                    location=f"sheet='{sheet_name}', column='{col}', quarter={first_q}",
                )
            df[col] = converted

        df = df.rename(columns=col_to_key)

        return cls(df, macro_variables=canonical_keys)

    @classmethod
    def from_workbook(
        cls,
        path: str | Path,
        config: dict,
    ) -> "EconomistProjections":
        """Read projections from an Excel workbook.

        The workbook must have a sheet named per
        ``config["conditioning"]["workbook_sheet"]`` (default
        ``"projections"``). The first column must be ``Quarter``; other
        columns must match the human-readable headers in
        ``config["data"]["macro_variables"][*]["column"]``.
        """
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(
                f"Projections workbook not found at {path}. "
                f"Generate one via: python -m cli.run_forecast "
                f"--generate-projections-template ..."
            )

        sheet = config.get("conditioning", {}).get("workbook_sheet", "projections")
        try:
            df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
        except ValueError as exc:
            raise InputValidationError(
                "missing_sheet",
                f"Projections workbook missing sheet '{sheet}': {exc}",
            ) from exc

        return cls.from_dataframe(df, config, sheet_name=sheet)


# ---------------------------------------------------------------------------
# ConditionalForecaster
# ---------------------------------------------------------------------------


class ConditionalForecaster:
    """Conditional forecasting via BGL-2015 / Waggoner-Zha-1999 (BVAR-SV adaptation)."""

    def __init__(
        self,
        results: BVARSVResults,
        projections: EconomistProjections,
        config_section: dict[str, Any],
        seed: int | None = None,
    ):
        self.results = results
        self.projections = projections
        self.config = dict(config_section)
        self.seed = seed

        self.variable_names = list(results.variable_names)
        self.macro_variables = list(config_section.get("macro_variables", []))
        if not self.macro_variables:
            raise ValueError(
                "config['conditioning']['macro_variables'] must not be empty"
            )

        # Validate that every config macro is in the panel, and resolve indices.
        try:
            self.macro_idx = np.array(
                [self.variable_names.index(name) for name in self.macro_variables],
                dtype=int,
            )
        except ValueError as exc:
            raise ValueError(
                f"macro_variables {self.macro_variables} not all in panel "
                f"{self.variable_names}: {exc}"
            ) from exc

        all_idx = set(range(len(self.variable_names)))
        target_idx_set = sorted(all_idx - set(self.macro_idx.tolist()))
        self.target_idx = np.array(target_idx_set, dtype=int)
        self.target_names = [self.variable_names[i] for i in self.target_idx]

        # Validate projections against the panel.
        self.projections.validate(results.data_used, results.n_lags)

        # Pre-extract projection array in canonical macro_variables order.
        self.projection_array = self.projections.to_array(self.macro_variables)
        if self.projection_array.shape[0] < int(self.config.get("horizon", 0) or 0):
            raise InputValidationError(
                "projection_horizon_short",
                f"projection horizon ({self.projection_array.shape[0]}) is shorter "
                f"than configured ({self.config['horizon']}); "
                "supply more projection rows or lower config.conditioning.horizon",
            )

    def forecast(
        self,
        n_paths_per_draw: int | None = None,
        n_draws_subsample: int | None = None,
    ) -> "ConditionalForecast":
        if n_paths_per_draw is None:
            n_paths_per_draw = int(self.config.get("n_paths_per_draw", 50))
        if n_draws_subsample is None:
            cfg_value = self.config.get("n_draws_subsample", None)
            n_draws_subsample = (
                int(cfg_value) if cfg_value is not None else None
            )

        # S0.3 — parameter guards.
        if n_paths_per_draw < 1:
            raise ValueError(
                f"n_paths_per_draw must be >= 1; got {n_paths_per_draw}"
            )
        if n_draws_subsample is not None and n_draws_subsample < 1:
            raise ValueError(
                f"n_draws_subsample must be >= 1 or None; got {n_draws_subsample}"
            )

        horizon = int(self.config.get("horizon", self.projections.horizon))
        if horizon <= 0:
            raise ValueError("conditioning.horizon must be positive")
        if horizon > self.projections.horizon:
            raise InputValidationError(
                "projection_horizon_short",
                f"horizon={horizon} exceeds projection rows={self.projections.horizon}",
            )

        strict = bool(self.config.get("enforce_strict_match", True))
        if not strict:
            unc = self.config.get("projection_uncertainty", {})
            stds = np.array(
                [float(unc.get(name, 0.0)) for name in self.macro_variables],
                dtype=float,
            )
            if np.any(stds <= 0):
                raise ValueError(
                    "projection_uncertainty must be positive for every macro "
                    "variable when enforce_strict_match=False; got "
                    f"{dict(zip(self.macro_variables, stds.tolist()))}"
                )
            # S0.3 Refinement 1 — tiered guard on projection_uncertainty.
            #   < 1e-4 : ValueError (almost certainly wrong)
            #   < 1e-2 : BVARWarning (borderline numerical zone)
            #   >= 1e-2: clean
            if np.any(stds < _HARD_MIN_PROJECTION_UNCERTAINTY):
                bad = {
                    name: float(s) for name, s in zip(self.macro_variables, stds)
                    if s < _HARD_MIN_PROJECTION_UNCERTAINTY
                }
                raise ValueError(
                    f"projection_uncertainty must be >= "
                    f"{_HARD_MIN_PROJECTION_UNCERTAINTY} for every macro "
                    f"variable when enforce_strict_match=False (Schur "
                    f"complement becomes singular below this threshold). "
                    f"Got: {bad}"
                )
            if np.any(stds < _SOFT_MIN_PROJECTION_UNCERTAINTY):
                from .exceptions import BVARWarning
                borderline = {
                    name: float(s) for name, s in zip(self.macro_variables, stds)
                    if s < _SOFT_MIN_PROJECTION_UNCERTAINTY
                }
                warnings.warn(
                    f"projection_uncertainty values below "
                    f"{_SOFT_MIN_PROJECTION_UNCERTAINTY} may produce "
                    f"numerically unstable conditional covariances on "
                    f"certain covariance shapes; verify Cholesky "
                    f"stability if using this configuration. Borderline "
                    f"values: {borderline}",
                    BVARWarning,
                    stacklevel=2,
                )
            R = np.diag(stds ** 2)
        else:
            R = None

        rng = np.random.default_rng(self.seed)
        n_kept = self.results.coefficients.shape[0]
        if n_draws_subsample is not None and n_draws_subsample < n_kept:
            draw_indices = rng.choice(n_kept, size=n_draws_subsample, replace=False)
            draw_indices.sort()
        else:
            draw_indices = np.arange(n_kept)

        n_draws_used = len(draw_indices)
        n_total = n_draws_used * n_paths_per_draw
        n_target = len(self.target_idx)
        n_macro = len(self.macro_idx)
        n_vars = len(self.variable_names)

        target_paths = np.empty((n_total, horizon, n_target))
        macro_paths = np.empty((n_total, horizon, n_macro))

        Y_full = self.results.data_used.to_numpy(dtype=float)
        last_obs_full = Y_full[-self.results.n_lags:].astype(float).copy()

        # Pre-draw all standard normals in deterministic C-order. The bulk
        # array packs target + macro + h slots per (path, t); the macro slot
        # is reserved even in strict mode so RNG consumption is uniform
        # across modes. The numba inner loop (Cholesky-based) consumes
        # these in the same order; same seed -> same output.
        all_z = rng.standard_normal((n_total, horizon, n_target + n_macro + n_vars))
        rng_z_target = all_z[:, :, :n_target]
        rng_z_macro = all_z[:, :, n_target:n_target + n_macro]
        rng_z_h = all_z[:, :, n_target + n_macro:]

        target_idx_arr = np.asarray(self.target_idx, dtype=np.int64)
        macro_idx_arr = np.asarray(self.macro_idx, dtype=np.int64)
        R_diag = (
            np.zeros(n_macro, dtype=float) if R is None
            else np.diag(R).astype(float).copy()
        )
        projection_arr = self.projection_array.astype(float)

        idx = 0
        for d in draw_indices:
            B = self.results.coefficients[d].astype(float)
            A = self.results.A_lower_triangular[d]
            try:
                A_inv = np.linalg.inv(A).astype(float)
            except np.linalg.LinAlgError:
                A_inv = np.linalg.pinv(A).astype(float)
            h_T_d = self.results.log_volatilities[d, -1, :].astype(float).copy()
            phi_d = self.results.phi[d].astype(float)
            omega_d = self.results.omega[d].astype(float)
            mu_d = self.results.mu[d].astype(float)

            z_t = rng_z_target[idx:idx + n_paths_per_draw].astype(float, copy=False)
            z_m = rng_z_macro[idx:idx + n_paths_per_draw].astype(float, copy=False)
            z_h = rng_z_h[idx:idx + n_paths_per_draw].astype(float, copy=False)

            tgt_d, mac_d = conditional_forecast_inner_loop(
                B, A_inv, h_T_d,
                omega_d, phi_d, mu_d,
                last_obs_full, projection_arr,
                target_idx_arr, macro_idx_arr,
                bool(R is None),  # strict
                R_diag,
                np.ascontiguousarray(z_t),
                np.ascontiguousarray(z_m),
                np.ascontiguousarray(z_h),
            )
            target_paths[idx:idx + n_paths_per_draw] = tgt_d
            macro_paths[idx:idx + n_paths_per_draw] = mac_d
            idx += n_paths_per_draw

        future_periods = pd.period_range(
            self.results.last_observation_period + 1, periods=horizon, freq="Q-DEC"
        )

        posterior_metadata = PosteriorMetadata(
            n_draws_used=int(n_draws_used),
            n_paths_per_draw=int(n_paths_per_draw),
            n_total_paths=int(n_total),
            horizon=int(horizon),
            enforce_strict_match=bool(strict),
            projection_uncertainty=(
                dict(self.config.get("projection_uncertainty", {})) if not strict else None
            ),
            macro_variables=list(self.macro_variables),
            target_variables=list(self.target_names),
            seed=self.seed,
            algorithm=(
                "BGL-2015 / Waggoner-Zha-1999 conditional Gaussian, "
                "per-time-t under BVAR-SV (path-dependent covariance)."
            ),
        )

        projections_used = self.projections.projections.iloc[:horizon].copy()

        return ConditionalForecast(
            target_paths=target_paths,
            target_names=list(self.target_names),
            macro_paths=macro_paths,
            macro_names=list(self.macro_variables),
            origin_period=self.results.last_observation_period,
            horizon=int(horizon),
            future_periods=future_periods,
            projections_used=projections_used,
            posterior_metadata=posterior_metadata,
        )

    # ------------------------------------------------------------------
    # Internal: per-draw, per-path simulation
    # ------------------------------------------------------------------

    def _simulate_path(
        self,
        B: np.ndarray,
        A_inv: np.ndarray,
        h_T: np.ndarray,
        phi: np.ndarray,
        omega: np.ndarray,
        mu: np.ndarray,
        last_obs: np.ndarray,
        R: np.ndarray | None,
        horizon: int,
        rng: np.random.Generator,
    ) -> tuple[np.ndarray, np.ndarray]:
        n_vars = len(self.variable_names)
        n_t = len(self.target_idx)
        n_m = len(self.macro_idx)
        target_idx = self.target_idx
        macro_idx = self.macro_idx

        state = last_obs.copy()
        h = h_T.copy()

        target_path = np.empty((horizon, n_t))
        macro_path = np.empty((horizon, n_m))

        for t in range(horizon):
            x = np.concatenate(([1.0], state[::-1].ravel()))
            mu_y = B @ x

            D = np.exp(h)  # (n_vars,)
            Sigma_y = (A_inv * D) @ A_inv.T  # equivalent to A_inv @ diag(D) @ A_inv.T

            mu_t = mu_y[target_idx]
            mu_m = mu_y[macro_idx]
            S_tt = Sigma_y[np.ix_(target_idx, target_idx)]
            S_tm = Sigma_y[np.ix_(target_idx, macro_idx)]
            S_mm = Sigma_y[np.ix_(macro_idx, macro_idx)]

            proj_t = self.projection_array[t]

            if R is None:
                # Strict match: macro_t = projection_t exactly.
                macro_t = proj_t
                residual = np.linalg.solve(S_mm, proj_t - mu_m)
                cond_mean_t = mu_t + S_tm @ residual
                cond_cov_t = S_tt - S_tm @ np.linalg.solve(S_mm, S_tm.T)
                target_t = rng.multivariate_normal(
                    cond_mean_t, _symmetrize(cond_cov_t)
                )
            else:
                # Soft match: projection_t = macro_t + eta, eta ~ N(0, R).
                S_pp = S_mm + R
                residual = np.linalg.solve(S_pp, proj_t - mu_m)
                cond_mean_t = mu_t + S_tm @ residual
                cond_mean_m = mu_m + S_mm @ residual
                cond_cov_t = S_tt - S_tm @ np.linalg.solve(S_pp, S_tm.T)
                cond_cov_m = S_mm - S_mm @ np.linalg.solve(S_pp, S_mm)
                cond_cov_tm = S_tm - S_tm @ np.linalg.solve(S_pp, S_mm)

                joint_mean = np.concatenate([cond_mean_t, cond_mean_m])
                joint_cov = np.empty((n_t + n_m, n_t + n_m))
                joint_cov[:n_t, :n_t] = cond_cov_t
                joint_cov[n_t:, n_t:] = cond_cov_m
                joint_cov[:n_t, n_t:] = cond_cov_tm
                joint_cov[n_t:, :n_t] = cond_cov_tm.T
                sample = rng.multivariate_normal(
                    joint_mean, _symmetrize(joint_cov)
                )
                target_t = sample[:n_t]
                macro_t = sample[n_t:]

            target_path[t] = target_t
            macro_path[t] = macro_t

            # Reassemble y_t and advance the state.
            y_t = np.empty(n_vars)
            y_t[target_idx] = target_t
            y_t[macro_idx] = macro_t
            state = np.vstack([state[1:], y_t[None, :]])

            # Evolve volatility (timing matches posterior_predictive_unconditional).
            h = phi * h + (1.0 - phi) * mu + omega * rng.standard_normal(n_vars)

        return target_path, macro_path


# ---------------------------------------------------------------------------
# Typed metadata containers (S0.7) and ConditionalForecast / YieldCurveForecast
# ---------------------------------------------------------------------------


@dataclass
class PosteriorMetadata:
    """Typed metadata for ConditionalForecast / YieldCurveForecast.

    Replaces the previously-untyped dict whose keys were documented only
    in the dataclass docstring. Wrappers reading ``cf.posterior_metadata``
    now get attribute access (``cf.posterior_metadata.n_total_paths``)
    with runtime type-safe construction.
    """
    n_draws_used: int
    n_paths_per_draw: int
    n_total_paths: int
    horizon: int
    enforce_strict_match: bool
    projection_uncertainty: dict[str, float] | None
    macro_variables: list[str]
    target_variables: list[str]
    seed: int | None
    algorithm: str

    def to_dict(self) -> dict:
        """Serialize for JSON. Used by save()."""
        return {
            "n_draws_used": self.n_draws_used,
            "n_paths_per_draw": self.n_paths_per_draw,
            "n_total_paths": self.n_total_paths,
            "horizon": self.horizon,
            "enforce_strict_match": self.enforce_strict_match,
            "projection_uncertainty": (
                dict(self.projection_uncertainty)
                if self.projection_uncertainty is not None else None
            ),
            "macro_variables": list(self.macro_variables),
            "target_variables": list(self.target_variables),
            "seed": self.seed,
            "algorithm": self.algorithm,
        }

    @classmethod
    def from_dict(cls, d: dict) -> "PosteriorMetadata":
        """Construct from a JSON-loaded dict. Used by load()."""
        return cls(
            n_draws_used=int(d["n_draws_used"]),
            n_paths_per_draw=int(d["n_paths_per_draw"]),
            n_total_paths=int(d["n_total_paths"]),
            horizon=int(d["horizon"]),
            enforce_strict_match=bool(d["enforce_strict_match"]),
            projection_uncertainty=(
                dict(d["projection_uncertainty"])
                if d.get("projection_uncertainty") is not None else None
            ),
            macro_variables=list(d["macro_variables"]),
            target_variables=list(d["target_variables"]),
            seed=d.get("seed"),
            algorithm=str(d["algorithm"]),
        )


@dataclass
class ConditionalForecast:
    target_paths: np.ndarray  # (n_total, horizon, n_target)
    target_names: list[str]
    macro_paths: np.ndarray  # (n_total, horizon, n_macro)
    macro_names: list[str]
    origin_period: pd.Period
    horizon: int
    future_periods: pd.PeriodIndex
    projections_used: pd.DataFrame
    posterior_metadata: PosteriorMetadata

    def to_yield_space(self, pca_dict: dict[str, Any]) -> "YieldCurveForecast":
        """Map PC-space target paths to per-maturity yield paths.

        Validates that ``target_names`` matches ``pca_dict['component_names']``;
        otherwise the PC ordering is wrong and the loadings projection would
        be silently incorrect.
        """
        comp = list(pca_dict.get("component_names", []))
        if comp != self.target_names:
            raise ValueError(
                f"target_names {self.target_names} do not match "
                f"pca_dict['component_names'] {comp}; ordering mismatch would "
                "produce incorrect yield projections."
            )
        loadings = np.asarray(pca_dict["loadings"], dtype=float)  # (n_maturities, n_pc)
        mean = np.asarray(pca_dict["mean"], dtype=float)         # (n_maturities,)

        # target_paths (P, H, n_pc) @ loadings.T (n_pc, n_maturities) -> (P, H, n_maturities)
        yield_paths = self.target_paths @ loadings.T + mean
        return YieldCurveForecast(
            yield_paths=yield_paths,
            maturity_names=list(pca_dict["yield_names"]),
            origin_period=self.origin_period,
            horizon=self.horizon,
            future_periods=self.future_periods,
            projections_used=self.projections_used.copy(),
            posterior_metadata=PosteriorMetadata.from_dict(
                self.posterior_metadata.to_dict()
            ),
        )

    def summary_statistics(
        self, quantiles: tuple[float, ...] = (0.05, 0.16, 0.50, 0.84, 0.95),
    ) -> dict[str, Any]:
        target_q = np.quantile(self.target_paths, list(quantiles), axis=0)
        macro_q = np.quantile(self.macro_paths, list(quantiles), axis=0)
        return {
            "quantiles": list(quantiles),
            "target_quantiles": {
                name: target_q[:, :, j].tolist()
                for j, name in enumerate(self.target_names)
            },
            "macro_quantiles": {
                name: macro_q[:, :, j].tolist()
                for j, name in enumerate(self.macro_names)
            },
            "future_periods": [str(p) for p in self.future_periods],
        }

    def save(self, path: str | Path) -> None:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        meta = {
            "target_names": list(self.target_names),
            "macro_names": list(self.macro_names),
            "origin_period": str(self.origin_period),
            "horizon": int(self.horizon),
            "future_periods": [str(p) for p in self.future_periods],
            "posterior_metadata": _to_jsonable(self.posterior_metadata.to_dict()),
        }
        proj_index = np.array([str(p) for p in self.projections_used.index])
        proj_columns = np.array(list(self.projections_used.columns))
        np.savez_compressed(
            path,
            target_paths=self.target_paths,
            macro_paths=self.macro_paths,
            projections_values=self.projections_used.to_numpy(dtype=float),
            projections_index=proj_index,
            projections_columns=proj_columns,
            metadata=np.array(json.dumps(meta)),
        )
        logger.info("Saved ConditionalForecast to %s", path)

    @classmethod
    def load(cls, path: str | Path) -> "ConditionalForecast":
        with np.load(path, allow_pickle=False) as f:
            meta = json.loads(str(f["metadata"]))
            proj_idx = pd.PeriodIndex(
                [pd.Period(s, freq="Q-DEC") for s in f["projections_index"]],
                freq="Q-DEC",
            )
            projections = pd.DataFrame(
                f["projections_values"],
                index=proj_idx,
                columns=[str(c) for c in f["projections_columns"]],
            )
            future_periods = pd.PeriodIndex(
                [pd.Period(s, freq="Q-DEC") for s in meta["future_periods"]],
                freq="Q-DEC",
            )
            return cls(
                target_paths=f["target_paths"].copy(),
                target_names=list(meta["target_names"]),
                macro_paths=f["macro_paths"].copy(),
                macro_names=list(meta["macro_names"]),
                origin_period=pd.Period(meta["origin_period"], freq="Q-DEC"),
                horizon=int(meta["horizon"]),
                future_periods=future_periods,
                projections_used=projections,
                posterior_metadata=PosteriorMetadata.from_dict(
                    meta["posterior_metadata"]
                ),
            )


@dataclass
class YieldCurveForecast:
    yield_paths: np.ndarray  # (n_total, horizon, n_maturities)
    maturity_names: list[str]
    origin_period: pd.Period
    horizon: int
    future_periods: pd.PeriodIndex
    projections_used: pd.DataFrame
    posterior_metadata: PosteriorMetadata

    def summary_statistics(
        self, quantiles: tuple[float, ...] = (0.05, 0.16, 0.50, 0.84, 0.95),
    ) -> dict[str, Any]:
        q = np.quantile(self.yield_paths, list(quantiles), axis=0)
        return {
            "quantiles": list(quantiles),
            "yield_quantiles": {
                name: q[:, :, j].tolist()
                for j, name in enumerate(self.maturity_names)
            },
            "future_periods": [str(p) for p in self.future_periods],
        }

    def to_dataframe(self, quantile: float = 0.50) -> pd.DataFrame:
        """Per-quantile yield levels: rows = forecast quarters, columns = maturities."""
        q = np.quantile(self.yield_paths, quantile, axis=0)
        return pd.DataFrame(
            q, index=self.future_periods, columns=self.maturity_names
        )

    def summary_dataframe(
        self,
        quantiles: tuple[float, ...] = (0.05, 0.16, 0.50, 0.84, 0.95),
    ) -> pd.DataFrame:
        """Long-form summary with multi-index columns (maturity, quantile).

        Easy to dump to CSV and ingest in a desk PowerPoint or report.
        """
        q = np.quantile(self.yield_paths, list(quantiles), axis=0)
        # q shape: (n_quantiles, horizon, n_maturities)
        n_q, h, n_m = q.shape
        # Reshape to (horizon, n_maturities * n_quantiles) with a multi-index column.
        q_labels = [f"q{int(round(x*100)):02d}" for x in quantiles]
        cols = pd.MultiIndex.from_product(
            [self.maturity_names, q_labels], names=["maturity", "quantile"]
        )
        # Reorder so columns are (maturity, quantile) groups
        # values: for each h, for each maturity m, for each quantile i: q[i, h, m]
        flat = np.transpose(q, (1, 2, 0)).reshape(h, n_m * n_q)
        return pd.DataFrame(flat, index=self.future_periods, columns=cols)

    def save(self, path: str | Path) -> None:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        meta = {
            "maturity_names": list(self.maturity_names),
            "origin_period": str(self.origin_period),
            "horizon": int(self.horizon),
            "future_periods": [str(p) for p in self.future_periods],
            "posterior_metadata": _to_jsonable(self.posterior_metadata.to_dict()),
        }
        proj_idx = np.array([str(p) for p in self.projections_used.index])
        proj_cols = np.array(list(self.projections_used.columns))
        np.savez_compressed(
            path,
            yield_paths=self.yield_paths,
            projections_values=self.projections_used.to_numpy(dtype=float),
            projections_index=proj_idx,
            projections_columns=proj_cols,
            metadata=np.array(json.dumps(meta)),
        )
        logger.info("Saved YieldCurveForecast to %s", path)

    @classmethod
    def load(cls, path: str | Path) -> "YieldCurveForecast":
        with np.load(path, allow_pickle=False) as f:
            meta = json.loads(str(f["metadata"]))
            proj_idx = pd.PeriodIndex(
                [pd.Period(s, freq="Q-DEC") for s in f["projections_index"]],
                freq="Q-DEC",
            )
            projections = pd.DataFrame(
                f["projections_values"],
                index=proj_idx,
                columns=[str(c) for c in f["projections_columns"]],
            )
            future_periods = pd.PeriodIndex(
                [pd.Period(s, freq="Q-DEC") for s in meta["future_periods"]],
                freq="Q-DEC",
            )
            return cls(
                yield_paths=f["yield_paths"].copy(),
                maturity_names=list(meta["maturity_names"]),
                origin_period=pd.Period(meta["origin_period"], freq="Q-DEC"),
                horizon=int(meta["horizon"]),
                future_periods=future_periods,
                projections_used=projections,
                posterior_metadata=PosteriorMetadata.from_dict(
                    meta["posterior_metadata"]
                ),
            )


# ---------------------------------------------------------------------------
# Projections workbook template generator
# ---------------------------------------------------------------------------


_PROJECTIONS_README_LINES: list[tuple[str, bool]] = [
    ("BVAR Yield Forecaster — Economist Projections Workbook", True),
    ("", False),
    (
        "This workbook supplies forward-looking macro projections to the "
        "conditional forecasting pipeline. The pipeline reads one sheet, "
        "named 'projections', and produces yield-curve forecasts conditional "
        "on the values you provide here.",
        False,
    ),
    ("", False),
    ("REQUIRED SHEETS", True),
    ("- projections : macro projections, one column per variable.", False),
    ("- README      : this sheet — ignored by the pipeline.", False),
    ("", False),
    ("REQUIRED COLUMNS — projections sheet", True),
    ("- Quarter                                       (YYYY-Q# format)", False),
    ("- Real GDP Growth (Q/Q SAAR)                    (percent)", False),
    ("- Headline CPI Inflation (Q/Q annualized)       (percent)", False),
    ("- Fed Funds Rate (quarterly average)            (percent)", False),
    ("", False),
    ("VALIDATION RULES (the pipeline enforces all of these)", True),
    ("1. Sheet 'projections' present, with all required columns.", False),
    ("2. First column is named 'Quarter'.", False),
    ("3. Quarter values match the regex ^\\d{4}-Q[1-4]$ (e.g. 2026-Q1).", False),
    ("4. Quarters are contiguous and start IMMEDIATELY after the panel's", False),
    ("   last observation. No gaps, no overlaps with the in-sample period.", False),
    ("5. No NaN values in any cell.", False),
    ("6. Each value is in [-50, 50] (rejects unit-scale errors).", False),
    ("", False),
    ("UNITS", True),
    ("All values are in PERCENT, not decimal form. Example:", False),
    ("    4.50  for a 4.50% Fed Funds Rate — correct", False),
    ("    0.045 for 4.50% — wrong (the pipeline will reject it).", False),
    ("", False),
    ("HORIZON", True),
    ("The number of rows you fill in determines the conditional-forecast", False),
    ("horizon. Default is 8 quarters (two years ahead). The horizon must be", False),
    ("at least as large as config.conditioning.horizon.", False),
    ("", False),
    ("USAGE", True),
    ("After filling in this workbook, run:", False),
    ("    python -m cli.run_forecast --forecast \\", False),
    ("        --config config/default.yaml \\", False),
    ("        --input data/raw/macro_yield_inputs.xlsx \\", False),
    ("        --projections <this-workbook.xlsx>", False),
]


def write_projections_template(
    output_path: str | Path,
    last_observation_period: pd.Period,
    horizon: int,
    macro_columns: list[str],
) -> Path:
    """Write a projections workbook template the desk can fill in.

    Parameters
    ----------
    output_path
        Path to write the .xlsx file.
    last_observation_period
        The panel's last observation; the template's first quarter is
        ``last_observation_period + 1``.
    horizon
        Number of quarters of projections to scaffold.
    macro_columns
        Human-readable Excel header names for the macro variables, in the
        order they should appear on the projections sheet.

    Returns
    -------
    Path
        The path written, for convenience.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "projections"

    headers = ["Quarter"] + list(macro_columns)
    ws.append(headers)

    quarters = pd.period_range(
        last_observation_period + 1, periods=horizon, freq="Q-DEC"
    )
    for q in quarters:
        # Format as YYYY-Q# to match the validation regex (str(Period) drops the dash).
        label = f"{q.year}-Q{q.quarter}"
        ws.append([label] + ["" for _ in macro_columns])

    ws_readme = wb.create_sheet("README")
    bold = Font(bold=True, size=12)
    wrap = Alignment(wrap_text=True, vertical="top")
    for row_idx, (text, is_bold) in enumerate(_PROJECTIONS_README_LINES, start=1):
        cell = ws_readme.cell(row=row_idx, column=1, value=text)
        cell.alignment = wrap
        if is_bold:
            cell.font = bold
    ws_readme.column_dimensions["A"].width = 90

    wb.save(output_path)
    logger.info(
        "Wrote projections template to %s (%d quarters: %s..%s)",
        output_path, horizon, quarters[0], quarters[-1],
    )
    return output_path


def apply_conditional_projection(*args, **kwargs):
    """Backward-compatible name for callers that imported the stub.

    Use ``ConditionalForecaster(...).forecast()`` instead.
    """
    raise NotImplementedError(
        "Use ConditionalForecaster(results, projections, config_section, seed) "
        ".forecast() instead."
    )
