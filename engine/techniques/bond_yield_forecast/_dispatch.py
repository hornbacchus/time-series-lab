"""Bond Yield Forecast — TSL technique dispatch entry point.

Implements ``run(ctx, progress_callback) -> dict`` per TSL technique
module convention (P-1 §8.1). Wraps the
``techniques.bond_yield_forecast`` subpackage (migrated from
``bvar-yield-forecaster`` v1.0.0-session-0-complete in Bond Yield
Forecast Session 1) behind the standard TSL RunContext / RunResponse
contract.

Two key deviations from typical TSL wrappers, justified by the Bond
Yield Forecast Integration Plan:

1.  **Workbook-input contract.** Most TSL wrappers consume
    ``ctx.series`` (Excel cell selection serialized as
    ``[{"name", "values"}]``). Bond Yield Forecast instead consumes a
    **multi-sheet xlsx workbook** referenced by
    ``ctx.params["input_workbook"]``. The workbook contains 3+ data
    sheets (BondYield_Macro, BondYield_Yields, BondYield_Projections).
    This deviation is endorsed by the integration plan §1.2 ("three-
    sheet bundled workbook") and reflects the wrapper's domain reality:
    the conditioning-on-projections architecture requires per-scenario
    projection sheets that don't fit a flat ``ctx.series`` payload.

2.  **No matplotlib in dispatch path.** Per integration plan §2.2,
    the wrapper does NOT invoke matplotlib. Charts come from
    Excel-native chart insertion at the C# add-in side (Session 3
    scope), not matplotlib png generation. The legacy CLI's
    ``_save_forecast_plots`` helper is in
    ``_legacy_cli.py.archive`` (does not migrate; archived).

Existing ``engine/techniques/bvar.py`` is the small Phase 1/2 BVAR
IRF/FEVD wrapper for ``technique_id="1c_bvar_irf_fevd"`` and is
unrelated. The two coexist; this module registers under
``technique_id="bond_yield_forecast"``.

See ``docs/bond_yield_forecast_integration/session_2_findings.md``
for design rationale and the friction-points §1 pre-flight
validation layer.
"""

from __future__ import annotations

import logging
import time
import traceback
import warnings
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from techniques.base import (
    RunContext,
    make_response,
    make_error_response,
    make_table,
)


# ---------------------------------------------------------------------------
# Catalog parameter bounds (mirror Session 0 ValueError thresholds + plan §2.3
# tightening). Used by the pre-flight validation layer.
# ---------------------------------------------------------------------------

# (min_inclusive, max_inclusive_or_None)
_PARAM_BOUNDS: dict[str, tuple[float | None, float | None]] = {
    # Minnesota prior tightness (Session 0 ValueError-on-zero guards)
    "lambda_1": (0.001, 2.0),
    "lambda_2": (0.001, None),
    "lambda_3": (0.001, None),
    "lambda_sc": (0.0, None),  # Session 0 allows 0 (loose-prior limit test)
    "lambda_io": (0.0, None),  # Session 0 allows 0
    # Conditional-forecast knobs (friction-points §3 OOM mitigation)
    "n_draws_subsample": (100, 5000),  # plan §2.3 tightened from 7000
    "n_paths_per_draw": (10, 500),
    "projection_uncertainty": (0.01, None),  # Session 0 tiered (≥0.01 clean)
    # Estimation chain
    "n_draws": (1000, 50000),
    "n_burn": (100, 20000),
    # Forecast horizon + scenario seed
    "horizon": (1, 20),
    "seed": (0, None),
}


# ---------------------------------------------------------------------------
# Pre-flight validation (friction-points §1 mitigation per plan §2.1.1)
# ---------------------------------------------------------------------------


class _PreflightFailure(Exception):
    """Raised when pre-flight validation rejects ctx.

    Caught by ``run()`` and mapped to a clean ``make_error_response``
    with structured fix suggestions, BEFORE invoking the deep BVAR
    stack (which would raise less actionable errors from inside
    ``BVARSV.__init__`` / ``ConditionalForecaster.__init__`` /
    ``MinnesotaPrior.dummy_observations`` etc.).
    """

    def __init__(self, message: str, fixes: list[str] | None = None) -> None:
        super().__init__(message)
        self.message = message
        self.fixes = fixes or []


def _preflight_validate_params(params: dict) -> list[str]:
    """Return list of human-readable validation errors (empty if OK).

    Checks every catalog-declared parameter against its bounds. The
    catalog also enforces these in the Task Pane UI (Session 3),
    but defense-in-depth: also reject here so a malformed dispatch
    JSON (e.g., from a future scripted invocation, or a stale UI
    bypassing newer catalog bounds) doesn't crash deep inside BVAR.
    """
    errors: list[str] = []
    for name, (lo, hi) in _PARAM_BOUNDS.items():
        if name not in params:
            continue  # catalog defaults apply downstream
        try:
            v = float(params[name])
        except (TypeError, ValueError):
            errors.append(
                f"Parameter '{name}': value {params[name]!r} is not a number."
            )
            continue
        if lo is not None and v < lo:
            errors.append(
                f"Parameter '{name}'={v} is below minimum {lo}. "
                f"This guards against Session 0-validated failure modes "
                f"(zero-divides, OOM, silent posterior corruption)."
            )
        if hi is not None and v > hi:
            errors.append(
                f"Parameter '{name}'={v} exceeds maximum {hi}. "
                f"This guards against documented OOM (n_draws_subsample) "
                f"and excessive runtime."
            )
    return errors


def _preflight_validate_workbook(workbook_path: Path, scenario: str) -> list[str]:
    """Verify the input workbook exists and is a readable .xlsx file.

    Pre-flight checks file-existence + .xlsx-loadability + minimum
    sheet count only. Deeper sheet-name validation is config-aware
    (sheet names come from ``config["data"]["macro_sheet"]``,
    ``config["data"]["unified_input"]["sheet_names"][...]``, etc.) and
    is deferred to ``read_unified_workbook``, which raises
    ``InputValidationError`` with sheet-specific messages. Those are
    caught by the dispatch's outer try/except and surfaced as a clean
    failure response.

    The integration plan §3.1 SAMPLE INPUT TEMPLATE uses fixed sheet
    names (``BondYield_Macro``, ``BondYield_Yields``,
    ``BondYield_Projections``); the canonical pre-S3 fixture uses
    BVAR-config-default names (``macro``, ``yields``,
    ``projections_baseline``). Both work; the config drives sheet
    resolution at the ``read_unified_workbook`` boundary.
    """
    errors: list[str] = []
    if not workbook_path.exists():
        errors.append(
            f"Input workbook not found at {workbook_path}. Use the "
            f"'Open Input Template' Ribbon menu item (Session 3) to "
            f"create a starter workbook."
        )
        return errors
    # Cheap loadability check (catches non-xlsx / corrupted files early).
    try:
        import openpyxl  # type: ignore

        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        n_sheets = len(wb.sheetnames)
        wb.close()
    except Exception as e:
        errors.append(
            f"Unable to open workbook at {workbook_path} as an .xlsx file: "
            f"{e.__class__.__name__}: {e}"
        )
        return errors

    if n_sheets < 2:
        errors.append(
            f"Workbook at {workbook_path} contains only {n_sheets} sheet(s); "
            f"Bond Yield Forecast requires at minimum a macro-history sheet, "
            f"a yield-history sheet, and a projections sheet. Sheet names are "
            f"resolved against config['data']['macro_sheet'], "
            f"config['data']['yield_sheet'], and "
            f"config['data']['unified_input']['sheet_names']['projections_<scenario>']."
        )
    return errors


def _preflight_validate(ctx: RunContext) -> list[str]:
    """Combined pre-flight: workbook + parameter bounds.

    Returns list of error messages; empty means OK.
    """
    errors: list[str] = []

    workbook = ctx.get_param("input_workbook")
    if not workbook:
        errors.append(
            "Missing required parameter 'input_workbook' (path to a "
            "3-sheet Bond Yield Forecast .xlsx workbook). Click "
            "'Open Input Template' on the Bond Yield Forecast Ribbon "
            "menu to create a starter workbook."
        )
        return errors  # can't validate sheets without a path

    scenario = str(ctx.get_param("scenario", "baseline") or "baseline")
    errors.extend(_preflight_validate_workbook(Path(workbook), scenario))
    errors.extend(_preflight_validate_params(ctx.params))
    return errors


# ---------------------------------------------------------------------------
# Panel-building (in-memory; bypasses build_panel's mandatory parquet write)
# ---------------------------------------------------------------------------


def _build_panel_in_memory(config: dict, raw: dict) -> dict[str, Any]:
    """Equivalent of ``build_panel(config, raw=...)`` without parquet write.

    Mirrors ``techniques.bond_yield_forecast.data.build_panel`` lines
    472-525 (validate_input → resolve_sample_window → align → fit_pca
    → project_to_pcs → concat) but skips the ``output_dir / *.parquet``
    write block, so pyarrow is not required as a TSL hard dependency.

    The TSL dispatch path uses the in-memory ``panel`` and ``pca_dict``
    directly; there is no consumer of the parquet outputs in TSL.
    """
    from techniques.bond_yield_forecast.data import (  # noqa: WPS433
        align_panel,
        fit_pca,
        project_to_pcs,
        validate_input,
        _resolve_sample_window,
    )

    clean = validate_input(raw, config)
    macro_df, yield_df = clean["macro"], clean["yields"]

    sample_start, sample_end = _resolve_sample_window(macro_df, yield_df, config)
    macro_df = macro_df.loc[sample_start:sample_end]
    yield_df = yield_df.loc[sample_start:sample_end]

    macro_cols = list(macro_df.columns)
    yield_cols = list(yield_df.columns)
    joined = pd.concat([macro_df, yield_df], axis=1)
    joined = align_panel(joined)
    macro_df = joined[macro_cols]
    yield_df = joined[yield_cols]

    pca_dict = fit_pca(
        yield_df,
        n_components=config["data"]["pca"]["n_components"],
        component_names=config["data"]["pca"]["component_names"],
    )
    pc_panel = project_to_pcs(yield_df, pca_dict)
    panel = pd.concat([macro_df, pc_panel], axis=1)

    return {
        "panel": panel,
        "yield_panel": yield_df,
        "pca": pca_dict,
        "sample": (sample_start, sample_end),
    }


# ---------------------------------------------------------------------------
# BVARWarning capture (plan §2.1.8)
# ---------------------------------------------------------------------------


def _categorize_warning(w: warnings.WarningMessage) -> str:
    """Return a Pattern J-style category string for a captured warning."""
    cls_name = w.category.__name__ if w.category else "Warning"
    return cls_name


def _summarize_warnings(captured: list[warnings.WarningMessage]) -> dict[str, Any]:
    """Aggregate captured warnings into audit-friendly summary."""
    from techniques.bond_yield_forecast.exceptions import BVARWarning  # noqa

    by_category: dict[str, int] = {}
    bvar_messages: list[str] = []
    other_messages: list[str] = []
    for w in captured:
        cat = _categorize_warning(w)
        by_category[cat] = by_category.get(cat, 0) + 1
        msg = str(w.message)
        if w.category and issubclass(w.category, BVARWarning):
            bvar_messages.append(f"[{cat}] {msg}")
        else:
            other_messages.append(f"[{cat}] {msg}")
    return {
        "by_category": by_category,
        "bvar_warning_messages": bvar_messages,
        "other_warning_messages": other_messages,
    }


# ---------------------------------------------------------------------------
# Output table construction
# ---------------------------------------------------------------------------


def _build_yield_forecast_table(yield_forecast, scenario: str) -> dict:
    """Yield-space forecast point estimate + 4 credible bands per maturity-horizon.

    ``yield_forecast.yield_paths`` shape (n_paths, horizon, n_maturities).
    Reduce across the n_paths dimension to median + 5/25/75/95 percentiles.
    Output one row per (horizon, maturity) cell.
    """
    paths = np.asarray(yield_forecast.yield_paths)
    horizon = int(paths.shape[1])
    n_mat = int(paths.shape[2])
    maturity_names = list(getattr(yield_forecast, "yield_names", [])) or [
        f"Maturity_{i}" for i in range(n_mat)
    ]
    proj_idx = list(getattr(yield_forecast, "projections_index", [])) or [
        f"H+{i + 1}" for i in range(horizon)
    ]

    columns = [
        "Scenario", "Horizon", "Maturity",
        "p05", "p25", "Median", "p75", "p95",
    ]
    rows = []
    for h in range(horizon):
        period_label = str(proj_idx[h]) if h < len(proj_idx) else f"H+{h + 1}"
        for m in range(n_mat):
            slab = paths[:, h, m]
            p05, p25, med, p75, p95 = np.percentile(
                slab, [5, 25, 50, 75, 95]
            )
            rows.append([
                scenario, period_label, str(maturity_names[m]),
                round(float(p05), 4), round(float(p25), 4),
                round(float(med), 4), round(float(p75), 4),
                round(float(p95), 4),
            ])
    return make_table("Yield Forecast (5/25/50/75/95 percentiles)", columns, rows)


def _build_macro_forecast_table(cond_forecast, scenario: str) -> dict:
    """Conditioning macro-path summary (median + 5/95 bands)."""
    paths = np.asarray(cond_forecast.macro_paths)
    horizon = int(paths.shape[1])
    n_mac = int(paths.shape[2])
    macro_names = list(getattr(cond_forecast, "projections_columns", [])) or [
        f"Macro_{i}" for i in range(n_mac)
    ]
    proj_idx = list(getattr(cond_forecast, "projections_index", [])) or [
        f"H+{i + 1}" for i in range(horizon)
    ]

    columns = ["Scenario", "Horizon", "Variable", "p05", "Median", "p95"]
    rows = []
    for h in range(horizon):
        period_label = str(proj_idx[h]) if h < len(proj_idx) else f"H+{h + 1}"
        for k in range(n_mac):
            slab = paths[:, h, k]
            p05, med, p95 = np.percentile(slab, [5, 50, 95])
            rows.append([
                scenario, period_label, str(macro_names[k]),
                round(float(p05), 4), round(float(med), 4),
                round(float(p95), 4),
            ])
    return make_table("Macro Conditioning Paths (5/50/95 percentiles)", columns, rows)


def _build_diagnostics_table(results) -> dict:
    """Per-equation BVAR-SV convergence diagnostics (ESS + Geweke z-scores)."""
    try:
        diag_df = results.convergence_diagnostics()
    except Exception:
        return make_table("Convergence Diagnostics", ["Note"],
                          [["Diagnostics computation failed; not surfaced."]])

    if isinstance(diag_df, pd.DataFrame):
        columns = ["Metric"] + list(diag_df.columns)
        rows = []
        for metric_name, row in diag_df.iterrows():
            row_vals = [str(metric_name)] + [
                round(float(v), 4) if isinstance(v, (int, float, np.floating))
                                     and not np.isnan(v) else str(v)
                for v in row.values
            ]
            rows.append(row_vals)
        return make_table("Convergence Diagnostics", columns, rows)
    return make_table("Convergence Diagnostics", ["Note"],
                      [["Diagnostics in unexpected format."]])


def _build_run_metadata_table(results, config: dict) -> dict:
    """Run-metadata audit row (n_draws, n_burn, runtime, etc.)."""
    pm = results.posterior_metadata
    rows = [
        ["Algorithm", "BVAR-SV (CCM-2019)"],
        ["n_variables", int(getattr(results, "n_variables",
                                    results.coefficients.shape[1]))],
        ["n_lags", int(config["model"]["lags"])],
        ["n_draws (kept)", int(results.coefficients.shape[0])],
        ["n_draws (total)", int(config["estimation"]["n_draws"])],
        ["n_burn", int(config["estimation"]["n_burn"])],
        ["seed", int(config["estimation"]["seed"])],
        ["runtime_seconds", round(float(getattr(pm, "runtime_seconds", 0.0)), 2)],
        ["commit_sha", str(getattr(pm, "commit_sha", "(none)"))],
    ]
    return make_table("Run Metadata", ["Field", "Value"], rows)


# ---------------------------------------------------------------------------
# Main run() entry point — TSL technique dispatch
# ---------------------------------------------------------------------------


def run(ctx: RunContext, progress_callback) -> dict:
    """Bond Yield Forecast dispatch.

    Parameters
    ----------
    ctx : RunContext
        Standard TSL run context. Must include
        ``ctx.params["input_workbook"]`` (absolute path to the 3-sheet
        Bond Yield Forecast .xlsx workbook). Optional:
        ``ctx.params["scenario"]`` (default ``"baseline"``); any
        catalog-declared parameter override (lambda_*, n_draws,
        n_burn, horizon, seed, projection_uncertainty,
        n_paths_per_draw, n_draws_subsample).
    progress_callback : callable
        TSL progress reporter ``(stage_label, percent[, message])``.

    Returns
    -------
    dict
        TSL ``RunResponse`` dict with tables, summary, audit_fields,
        warnings.
    """
    t0 = time.perf_counter()
    captured_warnings: list[warnings.WarningMessage] = []

    # -- Pre-flight validation (friction-points §1 mitigation) ----------
    progress_callback("Pre-flight validation", 5)
    errors = _preflight_validate(ctx)
    if errors:
        joined = "\n  - ".join(errors)
        return make_error_response(
            ctx,
            error_message=f"Pre-flight validation rejected the request:\n  - {joined}",
            error_fixes=[
                "Check the 'Open Input Template' Ribbon menu item to "
                "produce a workbook with the canonical 3-sheet format.",
                "Verify all parameter values are within the bounds "
                "documented in the catalog.",
                "If using a custom workbook, ensure the BondYield_Macro, "
                "BondYield_Yields, and BondYield_Projections sheets are "
                "present.",
            ],
        )

    # -- Imports deferred until after pre-flight (so pre-flight failures
    # don't pay the numba JIT compile cost or workbook-reader import cost)
    progress_callback("Loading subpackage", 8)
    try:
        from techniques.bond_yield_forecast.unified_input import (  # noqa
            read_unified_workbook,
        )
        from techniques.bond_yield_forecast.estimation import BVARSV
        from techniques.bond_yield_forecast.priors import MinnesotaPrior
        from techniques.bond_yield_forecast.conditioning import (
            ConditionalForecaster,
        )
        from techniques.bond_yield_forecast._paths import package_default_config
        from techniques.bond_yield_forecast.data import load_config
    except ImportError as e:
        return make_error_response(
            ctx,
            error_message=(
                f"Bond Yield Forecast subpackage import failed: "
                f"{e.__class__.__name__}: {e}. Ensure numba, pyyaml, "
                f"openpyxl are installed (engine/requirements.txt)."
            ),
        )

    # -- Capture all warnings under a single context manager so we can
    # surface them in audit_fields rather than letting BVARWarning
    # subclasses leak to stderr (plan §2.1.8).
    with warnings.catch_warnings(record=True) as captured_warnings:
        warnings.simplefilter("always")  # capture all categories

        try:
            # -- Load configuration (catalog params override defaults) -
            progress_callback("Loading configuration", 12)
            user_config = ctx.get_param("config_path", "") or ""
            config_path = Path(user_config) if user_config else package_default_config()
            config = load_config(config_path)
            config = _apply_param_overrides(config, ctx.params)
            scenario = str(ctx.get_param("scenario", "baseline") or "baseline")

            # -- Sheet-naming auto-detection (Session 3 carry-forward A) -
            # The Session 3 sample template uses BondYield_* sheet names
            # (per integration plan §3.1); the canonical pre-S3 fixture
            # uses BVAR-config-default names. Resolve which scheme applies
            # before invoking read_unified_workbook so a template-config
            # mismatch doesn't surface as an InputValidationError deep in
            # the reader.
            progress_callback("Detecting workbook sheet scheme", 15)
            workbook_path = Path(ctx.get_param("input_workbook"))
            config = _resolve_workbook_sheet_config(workbook_path, config, scenario)

            # -- Read unified workbook ---------------------------------
            progress_callback("Reading input workbook", 18)
            bundle = read_unified_workbook(workbook_path, scenario, config)

            # -- Build PCA panel (in-memory; no parquet write) ---------
            progress_callback("Building PCA panel", 25)
            panel_bundle = _build_panel_in_memory(config, bundle["raw"])
            panel = panel_bundle["panel"]
            pca_dict = panel_bundle["pca"]

            # -- BVAR-SV estimation (~20s on default fixture) ---------
            progress_callback(
                "Fitting BVAR-SV (this may take 15-30 seconds)", 35
            )
            variable_names = list(panel.columns)
            hp = config["model"]["hyperparameters"]["fixed"]
            prior = MinnesotaPrior(
                n_vars=len(variable_names),
                n_lags=int(config["model"]["lags"]),
                lambda_1=hp["lambda_1"], lambda_2=hp["lambda_2"],
                lambda_3=hp["lambda_3"], lambda_sc=hp["lambda_sc"],
                lambda_io=hp["lambda_io"],
                persistence_prior=config["model"]["persistence_prior"],
                variable_names=variable_names,
                training_data=panel,
            )
            bvar = BVARSV(
                data=panel,
                n_lags=int(config["model"]["lags"]),
                prior=prior,
                n_draws=int(config["estimation"]["n_draws"]),
                n_burn=int(config["estimation"]["n_burn"]),
                thinning=int(config["estimation"].get("thinning", 1)),
                seed=int(config["estimation"]["seed"]),
            )
            results = bvar.estimate()

            # -- Conditional forecast on the requested scenario --------
            progress_callback("Running conditional forecast", 75)
            forecaster = ConditionalForecaster(
                results=results,
                projections=bundle["projections"],
                config_section=config["conditioning"],
                seed=int(config["estimation"]["seed"]),
            )
            cf = forecaster.forecast()

            # -- Map PC-space forecast to yield space ------------------
            progress_callback("Projecting to yield space", 88)
            ycf = cf.to_yield_space(pca_dict)

            # -- Build output tables -----------------------------------
            progress_callback("Assembling output tables", 95)
            tables = [
                _build_yield_forecast_table(ycf, scenario),
                _build_macro_forecast_table(cf, scenario),
                _build_diagnostics_table(results),
                _build_run_metadata_table(results, config),
            ]

        except _PreflightFailure as exc:
            return make_error_response(
                ctx,
                error_message=exc.message,
                error_fixes=exc.fixes,
            )
        except Exception as exc:
            tb = traceback.format_exc(limit=6)
            return make_error_response(
                ctx,
                error_message=(
                    f"Bond Yield Forecast failed during execution: "
                    f"{exc.__class__.__name__}: {exc}\n"
                    f"Last frames:\n{tb}"
                ),
                error_fixes=[
                    "Check the engine log for full traceback.",
                    "If the failure is in BVAR-SV estimation, verify "
                    "n_draws / n_burn are within bounds and the workbook "
                    "history covers at least 60 quarters.",
                    "If the failure is in conditional forecasting, verify "
                    "the BondYield_Projections sheet's macro variable names "
                    "match BondYield_Macro's columns exactly (column-order "
                    "sensitive).",
                ],
            )

    # -- Aggregate captured warnings into audit_fields (plan §2.1.8) ----
    warnings_summary = _summarize_warnings(captured_warnings)
    plain_warnings = (
        warnings_summary["bvar_warning_messages"]
        + warnings_summary["other_warning_messages"]
    )

    # -- Plain-English summary ------------------------------------------
    n_kept = int(results.coefficients.shape[0])
    horizon = int(np.asarray(ycf.yield_paths).shape[1])
    summary = (
        f"Bond Yield Forecast: scenario='{scenario}'; "
        f"BVAR-SV estimated with {n_kept} kept draws (n_draws="
        f"{config['estimation']['n_draws']}, n_burn="
        f"{config['estimation']['n_burn']}). "
        f"Conditional forecast generated over {horizon}-quarter horizon "
        f"across {len(getattr(ycf, 'yield_names', [])) or 10} maturities "
        f"with 5/25/50/75/95 percentile bands. "
        f"Captured {len(captured_warnings)} warning(s) "
        f"({sum(warnings_summary['by_category'].values())} categorized)."
    )

    elapsed = time.perf_counter() - t0
    audit_fields = {
        "scenario": scenario,
        "n_lags": int(config["model"]["lags"]),
        "n_draws": int(config["estimation"]["n_draws"]),
        "n_burn": int(config["estimation"]["n_burn"]),
        "n_kept_draws": n_kept,
        "horizon": horizon,
        "n_paths_per_draw": int(config["conditioning"].get("n_paths_per_draw", 0)),
        "n_draws_subsample": config["conditioning"].get("n_draws_subsample"),
        "projection_uncertainty": _audit_safe(
            config["conditioning"].get("projection_uncertainty")
        ),
        "input_workbook": str(workbook_path),
        "wrapper_runtime_seconds": round(elapsed, 2),
        "warnings_count": len(captured_warnings),
        "warnings_by_category": warnings_summary["by_category"],
        "bvar_warning_messages": warnings_summary["bvar_warning_messages"],
    }

    return make_response(
        ctx,
        tables=tables,
        plain_english_summary=summary,
        warnings=plain_warnings,
        audit_fields=audit_fields,
        charting_suggestions=(
            "Recommended chart: line series of Median yield by Horizon, one "
            "line per Maturity, with shaded p25-p75 + p05-p95 bands. "
            "Plot in Excel via Insert → Line Chart on the Yield Forecast table."
        ),
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _resolve_workbook_sheet_config(
    workbook_path: Path, base_config: dict, scenario: str
) -> dict:
    """Detect sheet-naming scheme + return adjusted config.

    Two supported schemes (Session 3 / verification carry-forward A):

    1.  **Template scheme** (integration plan §3.1; Session 3 template):
        ``BondYield_Macro`` / ``BondYield_Yields`` / ``BondYield_Projections``.
        Rewrites ``config["data"]["macro_sheet"]`` etc. to match.

    2.  **Default scheme** (BVAR-config defaults; canonical pre-S3
        fixture): ``macro`` / ``yields`` / ``projections_baseline``
        etc. Returns ``base_config`` unchanged.

    Auto-detection avoids template-config mismatch as a runtime failure
    mode: a workbook-config drift would otherwise surface as a deep
    ``InputValidationError`` from inside ``read_unified_workbook``,
    which is harder to act on than an explicit "wrong sheet names" hint.

    Both schemes coexist; both work; this resolution is invisible to
    the caller.
    """
    import copy as _copy

    try:
        import openpyxl  # type: ignore

        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        sheets = set(wb.sheetnames)
        wb.close()
    except Exception:
        # Pre-flight already validated loadability; if this re-load fails,
        # fall back to base_config and let read_unified_workbook surface
        # the underlying issue.
        return base_config

    template_scheme_present = {
        "BondYield_Macro", "BondYield_Yields", "BondYield_Projections",
    }.issubset(sheets)

    if not template_scheme_present:
        # Default scheme (or some other variant); base config applies.
        return base_config

    cfg = _copy.deepcopy(base_config)
    # The unified-workbook reader resolves sheet names via
    # config["unified_input"]["sheet_names"] (see
    # techniques.bond_yield_forecast.unified_input._resolve_sheet_names).
    # Override macro_history / yields_history / projections_<scenario>
    # to point at the BondYield_* template sheet names.
    sn = cfg.setdefault("unified_input", {}).setdefault("sheet_names", {})
    sn["macro_history"] = "BondYield_Macro"
    sn["yields_history"] = "BondYield_Yields"
    sn[f"projections_{scenario}"] = "BondYield_Projections"

    # Some downstream paths (validate_input / build_panel for the legacy
    # two-file format) read config["data"]["macro_sheet"] /
    # ["yield_sheet"]. Mirror the override there too so re-entry through
    # those paths from the wrapper context resolves the right sheets.
    cfg.setdefault("data", {})["macro_sheet"] = "BondYield_Macro"
    cfg["data"]["yield_sheet"] = "BondYield_Yields"

    # Also map common alt-scenario aliases when present so a workbook
    # with additional BondYield_Projections_<scenario> sheets resolves.
    for alt in ("scenario_1", "scenario_2", "scenario_3", "scenario_4"):
        candidate = f"BondYield_Projections_{alt}"
        if candidate in sheets:
            sn[f"projections_{alt}"] = candidate
    return cfg


def _audit_safe(value):
    """Coerce config value to JSON-safe primitive for audit_fields.

    The config schema uses dict-shaped values for some keys (e.g.,
    ``projection_uncertainty`` is a per-variable dict in soft mode).
    Pass dicts/lists/None through unchanged; coerce numeric scalars
    to float for consistent JSON serialization.
    """
    if value is None or isinstance(value, (dict, list, bool, str)):
        return value
    try:
        return float(value)
    except (TypeError, ValueError):
        return str(value)


def _apply_param_overrides(config: dict, params: dict) -> dict:
    """Apply ctx.params overrides to the loaded config (in-place safe copy)."""
    import copy

    cfg = copy.deepcopy(config)

    def _set(path: list[str], key: str, source_key: str | None = None):
        sk = source_key or key
        if sk in params and params[sk] is not None:
            d = cfg
            for p in path[:-1]:
                d = d.setdefault(p, {})
            d[path[-1]] = params[sk]

    # Estimation knobs
    _set(["estimation", "n_draws"], "n_draws")
    _set(["estimation", "n_burn"], "n_burn")
    _set(["estimation", "seed"], "seed")
    # Model knobs
    _set(["model", "lags"], "lags")
    # Hyperparameter overrides (Minnesota lambdas)
    for k in ("lambda_1", "lambda_2", "lambda_3", "lambda_sc", "lambda_io"):
        if k in params and params[k] is not None:
            cfg.setdefault("model", {}).setdefault(
                "hyperparameters", {}
            ).setdefault("fixed", {})[k] = params[k]
    # Conditioning knobs
    _set(["conditioning", "n_paths_per_draw"], "n_paths_per_draw")
    _set(["conditioning", "n_draws_subsample"], "n_draws_subsample")
    _set(["conditioning", "projection_uncertainty"], "projection_uncertainty")
    _set(["conditioning", "horizon"], "horizon")
    # Honor TSL ctx.seed if estimation.seed not explicitly overridden
    if "seed" not in params and cfg.get("estimation", {}).get("seed") is None:
        cfg.setdefault("estimation", {})["seed"] = 42
    return cfg
