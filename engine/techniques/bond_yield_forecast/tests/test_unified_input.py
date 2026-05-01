"""Tests for src/bvar/unified_input.py — Step 5 unified-workbook architecture."""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from techniques.bond_yield_forecast.data import InputValidationError
from techniques.bond_yield_forecast.unified_input import (
    list_available_scenarios,
    read_unified_workbook,
    write_unified_template,
)


# ---------------------------------------------------------------------------
# Synthetic fixture builder
# ---------------------------------------------------------------------------


def _toy_config() -> dict:
    """Minimal config compatible with both the unified reader and the legacy
    data-pipeline helpers. Uses 4 macros (2 actual + 2 PCs) for compactness."""
    return {
        "data": {
            "input_file": "n/a",
            "macro_sheet": "macro",
            "yield_sheet": "yields",
            "macro_variables": {
                "real_gdp_growth": {
                    "column": "Real GDP Growth (Q/Q SAAR)",
                    "units": "percent",
                },
                "headline_cpi": {
                    "column": "Headline CPI Inflation (Q/Q annualized)",
                    "units": "percent",
                },
                "fed_funds_rate": {
                    "column": "Fed Funds Rate (quarterly average)",
                    "units": "percent",
                },
            },
            "yield_variables": {
                "treasury_3m": {"column": "3M", "maturity_years": 0.25},
                "treasury_6m": {"column": "6M", "maturity_years": 0.5},
                "treasury_1y": {"column": "1Y", "maturity_years": 1.0},
                "treasury_2y": {"column": "2Y", "maturity_years": 2.0},
                "treasury_3y": {"column": "3Y", "maturity_years": 3.0},
                "treasury_5y": {"column": "5Y", "maturity_years": 5.0},
                "treasury_7y": {"column": "7Y", "maturity_years": 7.0},
                "treasury_10y": {"column": "10Y", "maturity_years": 10.0},
                "treasury_20y": {"column": "20Y", "maturity_years": 20.0},
                "treasury_30y": {"column": "30Y", "maturity_years": 30.0},
            },
            "sample": {"start": "1990-Q1", "end": "latest"},
            "pca": {
                "n_components": 3,
                "component_names": ["pc1_level", "pc2_slope", "pc3_curvature"],
                "reporting_maturities": ["treasury_2y", "treasury_5y", "treasury_10y", "treasury_30y"],
            },
        },
        "model": {
            "lags": 1,
            "prior": "minnesota",
            "stochastic_volatility": True,
            "persistence_prior": {
                "real_gdp_growth": 0.0,
                "headline_cpi": 0.0,
                "fed_funds_rate": 1.0,
                "pc1_level": 1.0,
                "pc2_slope": 0.9,
                "pc3_curvature": 0.5,
            },
            "hyperparameters": {
                "method": "fixed",
                "fixed": {"lambda_1": 0.2, "lambda_2": 0.5, "lambda_3": 1.0,
                          "lambda_sc": 1.0, "lambda_io": 1.0},
            },
        },
        "estimation": {"n_draws": 600, "n_burn": 200, "thinning": 1, "seed": 17},
        "forecast": {"horizon_default": 4, "horizon_midyear": 2, "bands": [0.16, 0.5, 0.84]},
        "conditioning": {
            "horizon": 4,
            "n_paths_per_draw": 4,
            "n_draws_subsample": 100,
            "macro_variables": ["real_gdp_growth", "headline_cpi", "fed_funds_rate"],
            "enforce_strict_match": True,
            "projection_uncertainty": {
                "real_gdp_growth": 0.5, "headline_cpi": 0.3, "fed_funds_rate": 0.1,
            },
            "workbook_sheet": "projections_baseline",
        },
        "unified_input": {
            "workbook_filename": "bvar_inputs.xlsx",
            "sheet_names": {
                "macro_history": "macro",
                "yields_history": "yields",
                "projections_baseline": "projections_baseline",
                "projections_scenario_1": "projections_scenario_1",
                "projections_scenario_2": "projections_scenario_2",
                "projections_scenario_3": "projections_scenario_3",
                "projections_scenario_4": "projections_scenario_4",
                "readme": "README",
            },
            "default_scenario": "baseline",
        },
    }


def _build_synthetic_unified_workbook(
    output_path: Path,
    *,
    n_history_quarters: int = 60,
    horizon: int = 4,
    populated_scenarios: list[str] = ("baseline",),
    seed: int = 0,
) -> Path:
    """Build a synthetic 8-sheet unified workbook for tests.

    `populated_scenarios` lists which projection sheets to fill with data
    rows; the rest stay header-only. The history sheets always have
    `n_history_quarters` plausible synthetic rows.
    """
    rng = np.random.default_rng(seed)
    cfg = _toy_config()
    macro_columns = [s["column"] for s in cfg["data"]["macro_variables"].values()]
    yield_columns = [s["column"] for s in cfg["data"]["yield_variables"].values()]
    sheet_names = cfg["unified_input"]["sheet_names"]

    start = pd.Period("2010-Q1", freq="Q-DEC")
    history_idx = pd.period_range(start, periods=n_history_quarters, freq="Q-DEC")
    last_history = history_idx[-1]
    proj_idx = pd.period_range(last_history + 1, periods=horizon, freq="Q-DEC")

    # Plausible synthetic series (just persistent random walks within sane bounds).
    macro_data = np.column_stack([
        2.0 + rng.standard_normal(n_history_quarters) * 0.5,    # GDP growth
        2.0 + rng.standard_normal(n_history_quarters) * 0.4,    # CPI
        2.5 + np.cumsum(rng.standard_normal(n_history_quarters) * 0.1),  # FFR drift
    ])
    yield_data = np.zeros((n_history_quarters, len(yield_columns)))
    for i, m in enumerate(yield_columns):
        years = float({"3M": 0.25, "6M": 0.5, "1Y": 1.0, "2Y": 2.0, "3Y": 3.0,
                       "5Y": 5.0, "7Y": 7.0, "10Y": 10.0, "20Y": 20.0, "30Y": 30.0}[m])
        yield_data[:, i] = 3.0 + 0.5 * np.log(1 + years) + rng.standard_normal(n_history_quarters) * 0.2

    wb = Workbook()
    wb.remove(wb.active)

    # macro sheet
    ws = wb.create_sheet(sheet_names["macro_history"])
    ws.append(["Quarter"] + macro_columns)
    for i, q in enumerate(history_idx):
        ws.append([f"{q.year}-Q{q.quarter}"] + [round(float(v), 3) for v in macro_data[i]])

    # yields sheet
    ws = wb.create_sheet(sheet_names["yields_history"])
    ws.append(["Quarter"] + yield_columns)
    for i, q in enumerate(history_idx):
        ws.append([f"{q.year}-Q{q.quarter}"] + [round(float(v), 3) for v in yield_data[i]])

    # 5 projection sheets: header-only, then optionally populated.
    for key in ["projections_baseline", "projections_scenario_1",
                "projections_scenario_2", "projections_scenario_3",
                "projections_scenario_4"]:
        sheet = sheet_names[key]
        ws = wb.create_sheet(sheet)
        ws.append(["Quarter"] + macro_columns)
        scenario_key = key.replace("projections_", "")  # "baseline", "scenario_1", ...
        if scenario_key in populated_scenarios:
            proj_macro = np.column_stack([
                2.0 + rng.standard_normal(horizon) * 0.3,
                2.5 + rng.standard_normal(horizon) * 0.3,
                3.0 + rng.standard_normal(horizon) * 0.3,
            ])
            for i, q in enumerate(proj_idx):
                ws.append([f"{q.year}-Q{q.quarter}"] +
                          [round(float(v), 3) for v in proj_macro[i]])

    # README sheet
    ws_readme = wb.create_sheet(sheet_names["readme"])
    ws_readme.append(["Synthetic test workbook for tests/test_unified_input.py"])

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_read_unified_workbook_with_baseline_scenario(tmp_path):
    cfg = _toy_config()
    wb_path = tmp_path / "bvar_inputs.xlsx"
    _build_synthetic_unified_workbook(wb_path, populated_scenarios=["baseline"])

    bundle = read_unified_workbook(wb_path, "baseline", cfg)
    assert "macro_history" in bundle
    assert "yields_history" in bundle
    assert "projections" in bundle
    assert "raw" in bundle
    assert "metadata" in bundle

    assert isinstance(bundle["macro_history"].index, pd.PeriodIndex)
    assert list(bundle["macro_history"].columns) == [
        "real_gdp_growth", "headline_cpi", "fed_funds_rate"
    ]
    assert bundle["projections"].horizon == 4
    assert bundle["metadata"]["scenario"] == "baseline"
    assert bundle["metadata"]["available_scenarios"] == ["baseline"]


def test_read_unified_workbook_rejects_empty_scenario(tmp_path):
    cfg = _toy_config()
    wb_path = tmp_path / "bvar_inputs.xlsx"
    # Only baseline populated; scenario_3 is header-only.
    _build_synthetic_unified_workbook(wb_path, populated_scenarios=["baseline"])

    with pytest.raises(InputValidationError, match="empty_scenario") as exc_info:
        read_unified_workbook(wb_path, "scenario_3", cfg)
    # Verify message includes a remediation hint.
    assert "scenario_3" in str(exc_info.value)
    assert "Fill in the projections sheet" in str(exc_info.value)


def test_read_unified_workbook_validates_history_alignment(tmp_path):
    """Macro and yields sheets with different Quarter indexes must fail.

    Either rule may fire (whichever the validator hits first):
      - 'quarter_gap': yields has a missing quarter inside its span.
      - 'index_mismatch': macro and yields have different Quarter sets.
    Both indicate the same underlying problem; we accept either.
    """
    cfg = _toy_config()
    wb_path = tmp_path / "bvar_inputs.xlsx"
    _build_synthetic_unified_workbook(wb_path, populated_scenarios=["baseline"])

    # Corrupt the workbook: drop a row from yields, leaving macro intact.
    wb = load_workbook(wb_path)
    yields_ws = wb["yields"]
    yields_ws.delete_rows(5)  # drop one history row
    wb.save(wb_path)

    with pytest.raises(InputValidationError, match="quarter_gap|index_mismatch"):
        read_unified_workbook(wb_path, "baseline", cfg)


def test_read_unified_workbook_validates_projection_alignment(tmp_path):
    """Projection quarters must begin immediately after history's last quarter."""
    cfg = _toy_config()
    wb_path = tmp_path / "bvar_inputs.xlsx"
    _build_synthetic_unified_workbook(wb_path, populated_scenarios=["baseline"])

    # Corrupt the workbook: set baseline's first projection Quarter to two
    # quarters after history's last (creates a gap).
    wb = load_workbook(wb_path)
    macro_ws = wb["macro"]
    last_q = macro_ws.cell(row=macro_ws.max_row, column=1).value
    last_period = pd.Period(last_q, freq="Q-DEC")
    bad_start = last_period + 2  # gap
    bw = wb["projections_baseline"]
    bw.cell(row=2, column=1).value = f"{bad_start.year}-Q{bad_start.quarter}"
    # Adjust subsequent rows too so the projections are still contiguous internally
    # but offset by one quarter.
    for offset, r in enumerate(range(3, bw.max_row + 1), start=1):
        q = bad_start + offset
        bw.cell(row=r, column=1).value = f"{q.year}-Q{q.quarter}"
    wb.save(wb_path)

    with pytest.raises(InputValidationError, match="projection_alignment"):
        read_unified_workbook(wb_path, "baseline", cfg)


def test_write_unified_template_produces_8_sheets(tmp_path):
    cfg = _toy_config()
    out = tmp_path / "tmpl.xlsx"
    write_unified_template(output_path=out, horizon=8, config=cfg)

    wb = load_workbook(out)
    expected = {
        "macro", "yields",
        "projections_baseline", "projections_scenario_1",
        "projections_scenario_2", "projections_scenario_3",
        "projections_scenario_4",
        "README",
    }
    assert set(wb.sheetnames) == expected

    # Each projection sheet must have horizon rows + 1 header.
    for s in ("projections_baseline", "projections_scenario_1",
              "projections_scenario_2", "projections_scenario_3",
              "projections_scenario_4"):
        ws = wb[s]
        assert ws.max_row == 9  # 1 header + 8 quarter rows
        assert ws.cell(row=1, column=1).value == "Quarter"
        # Quarter labels match YYYY-Q# regex (with the dash).
        for r in range(2, 10):
            label = ws.cell(row=r, column=1).value
            assert isinstance(label, str)
            assert "-Q" in label


def test_write_unified_template_with_copy_history(tmp_path):
    """Migration helper: --copy-history-from copies macro/yields verbatim."""
    cfg = _toy_config()
    src = tmp_path / "old_inputs.xlsx"
    _build_synthetic_unified_workbook(src, n_history_quarters=20)

    dst = tmp_path / "new_inputs.xlsx"
    write_unified_template(
        output_path=dst, horizon=8, copy_history_from=src, config=cfg,
    )

    src_wb = load_workbook(src)
    dst_wb = load_workbook(dst)

    src_macro_rows = list(src_wb["macro"].iter_rows(values_only=True))
    dst_macro_rows = list(dst_wb["macro"].iter_rows(values_only=True))
    assert src_macro_rows == dst_macro_rows

    src_yields_rows = list(src_wb["yields"].iter_rows(values_only=True))
    dst_yields_rows = list(dst_wb["yields"].iter_rows(values_only=True))
    assert src_yields_rows == dst_yields_rows

    # Projection sheets in dst should be header-only with quarter labels
    # starting immediately after src's last history quarter.
    src_macro_last = src_macro_rows[-1][0]
    last_period = pd.Period(src_macro_last, freq="Q-DEC")
    expected_first_proj = last_period + 1
    actual_first_proj = pd.Period(
        dst_wb["projections_baseline"].cell(row=2, column=1).value, freq="Q-DEC"
    )
    assert actual_first_proj == expected_first_proj


def test_list_available_scenarios(tmp_path):
    cfg = _toy_config()
    wb_path = tmp_path / "bvar_inputs.xlsx"
    _build_synthetic_unified_workbook(
        wb_path, populated_scenarios=["baseline", "scenario_2"],
    )

    available = list_available_scenarios(wb_path, cfg)
    assert available == ["baseline", "scenario_2"]


@pytest.mark.skip(
    reason="CLI did not migrate to TSL (archived as _legacy_cli.py.archive); this test exercised CLI side-effects that are now reproduced via the TSL engine_worker dispatch path (Session 2+) instead."
)
def test_forecast_with_unified_input_end_to_end(tmp_path):
    """Full smoke: synthetic workbook → main(['--forecast']) → expected outputs."""
    from cli.run_forecast import main
    import yaml

    cfg = _toy_config()
    cfg_path = tmp_path / "config.yaml"
    cfg_path.write_text(yaml.safe_dump(cfg, sort_keys=False), encoding="utf-8")
    wb_path = tmp_path / "bvar_inputs.xlsx"
    _build_synthetic_unified_workbook(
        wb_path, populated_scenarios=["baseline"], n_history_quarters=80,
    )
    out_dir = tmp_path / "out"

    rc = main([
        "--forecast", "--config", str(cfg_path),
        "--input", str(wb_path), "--scenario", "baseline",
        "--output-dir", str(out_dir), "--no-confirm",
    ])
    assert rc == 0
    assert (out_dir / "estimation_results.npz").exists()
    assert (out_dir / "posterior_summary.txt").exists()
    assert (out_dir / "convergence_diagnostics.csv").exists()
    assert (out_dir / "baseline").is_dir()
    assert (out_dir / "baseline" / "conditional_forecast.npz").exists()
    assert (out_dir / "baseline" / "yield_curve_forecast.npz").exists()
    assert (out_dir / "baseline" / "forecast_summary.csv").exists()
    assert (out_dir / "baseline" / "forecast_plots").is_dir()


@pytest.mark.skip(
    reason="CLI did not migrate to TSL (archived as _legacy_cli.py.archive); this test exercised CLI side-effects that are now reproduced via the TSL engine_worker dispatch path (Session 2+) instead."
)
def test_forecast_scenario_all_produces_subdirs(tmp_path):
    """--scenario all on a workbook with baseline + 2 alternates produces three
    scenario subdirs sharing one estimation_results.npz."""
    from cli.run_forecast import main
    import yaml

    cfg = _toy_config()
    cfg_path = tmp_path / "config.yaml"
    cfg_path.write_text(yaml.safe_dump(cfg, sort_keys=False), encoding="utf-8")
    wb_path = tmp_path / "bvar_inputs.xlsx"
    _build_synthetic_unified_workbook(
        wb_path,
        populated_scenarios=["baseline", "scenario_1", "scenario_3"],
        n_history_quarters=80,
    )
    out_dir = tmp_path / "out_all"

    rc = main([
        "--forecast", "--config", str(cfg_path),
        "--input", str(wb_path), "--scenario", "all",
        "--output-dir", str(out_dir), "--no-confirm",
    ])
    assert rc == 0
    assert (out_dir / "estimation_results.npz").exists()
    assert (out_dir / "baseline" / "yield_curve_forecast.npz").exists()
    assert (out_dir / "scenario_1" / "yield_curve_forecast.npz").exists()
    assert (out_dir / "scenario_3" / "yield_curve_forecast.npz").exists()
    # Empty scenarios should NOT have subdirs.
    assert not (out_dir / "scenario_2").exists()
    assert not (out_dir / "scenario_4").exists()


@pytest.mark.skip(
    reason="CLI did not migrate to TSL (archived as _legacy_cli.py.archive); this test exercised CLI side-effects that are now reproduced via the TSL engine_worker dispatch path (Session 2+) instead."
)
def test_deprecated_command_emits_notice(tmp_path, capsys):
    """Running --process-data emits the Step-5 deprecation notice on stderr."""
    from cli.run_forecast import main
    import yaml

    cfg = _toy_config()
    cfg_path = tmp_path / "config.yaml"
    cfg_path.write_text(yaml.safe_dump(cfg, sort_keys=False), encoding="utf-8")
    wb_path = tmp_path / "bvar_inputs.xlsx"
    _build_synthetic_unified_workbook(wb_path, populated_scenarios=["baseline"])
    out_dir = tmp_path / "out"

    # --process-data uses macro/yields sheets — same names in unified workbook.
    cfg_with_path = dict(cfg)
    cfg_with_path["data"] = dict(cfg["data"])
    cfg_with_path["data"]["input_file"] = str(wb_path)
    cfg_path.write_text(yaml.safe_dump(cfg_with_path, sort_keys=False), encoding="utf-8")

    main([
        "--process-data", "--config", str(cfg_path),
        "--input", str(wb_path), "--output-dir", str(out_dir),
    ])
    captured = capsys.readouterr()
    assert "DEPRECATION NOTICE: --process-data is deprecated" in captured.err
