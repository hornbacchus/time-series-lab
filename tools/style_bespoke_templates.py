"""Apply the Bespoke input-cell convention (owner ruling, K3.2) to the
styled-in-place templates: BYF + Breakeven Payrolls.

CONVENTION: every cell the tool's READER consumes gets (a) solid yellow
#FFFF00 fill and (b) thin black borders on all four sides. The input set is
DERIVED from each reader, never eyeballed (the inert-lesson applied to
styling):

BYF (read_unified_workbook, unified_input.py -- pd.read_excel, header row 1,
data from row 2; README is display-only):
  * BondYield_Macro        rows 2..144 x cols 1..7
  * BondYield_Yields       rows 2..144 x cols 1..35
  * BondYield_Projections  rows 2..9   x cols 1..6
  (the pre-filled extents; users extending paste past them -- the reader
  follows the data, not the paint)
  NOTE: a write_unified_template generator EXISTS (unified_input.py l.374)
  but requires copy_history_from (the original source workbook, not in-repo)
  to reproduce the baked history -- the committed template is a PINNED-DATA
  artifact, so it is styled in place here; any future regeneration must
  re-apply the convention.

Breakeven (workbook_input.py -- header row 4 on the data tabs; _kv_scan on
the key-value tabs; cbo_annual / actuals_for_charts / control_vintage /
Instructions are NOT reader-read -> never styled; _meta is scanned but NO
value is consumed -> display-only per the read-but-ignored precedent
(Kronos T/top_p)):
  * cbo_quarterly       rows 5..356 x cols 1..7
  * population_monthly  rows 5..810 x cols 1..5 (incl. the CNP16OV append
    region; the separator row sits inside the scanned block)
  * scenario_inputs     the column-B value cell of the consumed key
    net_migration_all_ages
  * scalars             the column-B value cells of the 5 consumed keys
    (pop_base_v2025_dec2025, ces_se_monthly_nfp_change, share_16plus,
    census_base_pace, census_migration_ref)

GUARD: cell VALUES are snapshotted before styling and asserted byte-equal
after (styling-only delta; the file bytes change, the data must not).
openpyxl round-trip hazard check: neither template carries charts; the
Breakeven scenario_inputs data validations are an openpyxl-supported feature
and survive the round-trip.
"""

import sys

from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill, Side

FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_THIN = Side(style="thin", color="000000")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

BYF = ("engine/techniques/bond_yield_forecast/resources/templates/"
       "bond_yield_forecast_input_template.xlsx")
BRK = ("engine/techniques/breakeven_payroll/resources/templates/"
       "breakeven_payroll_input_template.xlsx")

BYF_RANGES = [  # (sheet, row_first, row_last, col_first, col_last)
    ("BondYield_Macro", 2, 144, 1, 7),
    ("BondYield_Yields", 2, 144, 1, 35),
    ("BondYield_Projections", 2, 9, 1, 6),
]
BRK_RANGES = [
    ("cbo_quarterly", 5, 356, 1, 7),
    ("population_monthly", 5, 810, 1, 5),
]
BRK_KV = {  # sheet -> the consumed keys (column-B value cells)
    "scenario_inputs": ["net_migration_all_ages"],
    "scalars": ["pop_base_v2025_dec2025", "ces_se_monthly_nfp_change",
                "share_16plus", "census_base_pace", "census_migration_ref"],
}


def snapshot(path):
    wb = load_workbook(path, read_only=True)
    out = {}
    for ws in wb.worksheets:
        out[ws.title] = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    return out


def style(path, ranges, kv=None):
    wb = load_workbook(path)
    n = 0
    for sheet, r0, r1, c0, c1 in ranges:
        ws = wb[sheet]
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = FILL
                cell.border = BORDER
                n += 1
    for sheet, keys in (kv or {}).items():
        ws = wb[sheet]
        for row in ws.iter_rows(min_col=1, max_col=2):
            key = row[0].value
            if isinstance(key, str) and key.strip() in keys:
                cell = ws.cell(row=row[0].row, column=2)
                cell.fill = FILL
                cell.border = BORDER
                n += 1
    wb.save(path)
    return n


def main():
    for name, path, ranges, kv in (("BYF", BYF, BYF_RANGES, None),
                                   ("Breakeven", BRK, BRK_RANGES, BRK_KV)):
        before = snapshot(path)
        n = style(path, ranges, kv)
        after = snapshot(path)
        if before != after:
            sys.exit(f"FATAL: {name} cell VALUES changed during styling -- aborting.")
        print(f"{name}: {n} input cells styled; value-diff = ZERO changes.")


if __name__ == "__main__":
    main()
