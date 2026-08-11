"""One-off template updater: breakeven_payroll_input_template.xlsx v1 (9 tabs)
-> v2 (10 tabs) for the preset-only re-port @ standalone 99c03ba (CFL-525).

Additions (append-only — NO row insertion/shifting, so every pre-existing cell
keeps its address and value):
  1. NEW tab `gross_migration_sums` — per year x status x flow x age-bucket
     {u16, 16plus} people sums at FULL float precision, baked from the
     STANDALONE repo's own groupby output (passed in as a CSV; no TSL-side
     re-summation of the 96k-row source). Age -1 counts in u16, so
     u16 + 16plus == all-ages exactly.
  2. `scenario_inputs` r16: the CBO Feb-2026 preset reference row (display
     block only — the active-scenario knob stays numeric per 2aa712f).
  3. `_meta`: re-port stamp keys (pin, date, grossMigration sha256, note).
  4. `control_vintage`: one documentary pointer row.

Post-edit assertions (amendment-2 contract): ALL cells of the 9 pre-existing
tabs value-identical EXCEPT the three enumerated appended ranges; the yellow
knob styling survives; tab count == 10; the sums tab reproduces the committed-
vintage tie-outs (2026: 1,535,765 / 961,806 / +573,959; 16+ net +499,825;
shares .8940/.9078/.8708 at <=1e-6).

Usage: python tools/update_breakeven_template_v2.py <gm_sums.csv>
(gm_sums.csv columns: year,immigration_status,migration_flow,bucket,people —
 produced in the standalone repo by the groupby documented in its
 PARITY_FIXTURE.md re-port section.)
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = (ROOT / "engine" / "techniques" / "breakeven_payroll" / "resources"
            / "templates" / "breakeven_payroll_input_template.xlsx")

PIN = "99c03ba"
REPORT_DATE = "2026-08-08"
GM_SHA256 = "d64868ad1b544854c2ff4a9064ddeb3755c08f136d4c519531406e6388a7b9a5"

# The three enumerated appended ranges (tab, {(row, col), ...}) — everything
# else pre-existing must be value-identical.
APPEND_SCENARIO_ROW = 16          # scenario_inputs A16:C16
APPEND_META_FIRST_ROW = 20        # _meta A20:C23 (4 keys)
_META_KEYS = [
    ("reported_pin", PIN, "preset-only re-port; source repo tip"),
    ("reported_date", REPORT_DATE, "re-port cut date (owner decision)"),
    ("gross_migration_sha256", GM_SHA256,
     "full-file hash of the source grossMigration CSV (committed vintage)"),
    ("vintage_note", "see standalone tests/reports/cfl525_extensions_audit.md",
     "the gross-flow vintage reissue (nets preserved <=37)"),
]


def snapshot(wb):
    snap = {}
    for name in wb.sheetnames:
        ws = wb[name]
        cells = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    cells[(c.row, c.column)] = c.value
        snap[name] = cells
    return snap


def main(sums_csv: Path) -> None:
    sums = pd.read_csv(sums_csv)
    assert list(sums.columns) == ["year", "immigration_status",
                                  "migration_flow", "bucket", "people"]

    wb = openpyxl.load_workbook(TEMPLATE)
    assert len(wb.sheetnames) == 9, wb.sheetnames
    before = snapshot(wb)

    # -- 1) the sums tab ----------------------------------------------------
    ws = wb.create_sheet("gross_migration_sums")
    ws["A1"] = "CBO gross migration — derived sums (CFL-525 re-port)"
    ws["A2"] = ("Baked from the standalone repo's groupby of the committed "
                f"pub-61879 file (sha256 {GM_SHA256}); age -1 counts in u16, "
                "so u16+16plus == all-ages exactly. Reference data — do not edit.")
    hdr = ["year", "immigration_status", "migration_flow", "bucket", "people"]
    for j, h in enumerate(hdr, start=1):
        ws.cell(row=4, column=j, value=h)
    for i, row in enumerate(sums.itertuples(index=False), start=5):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    # -- 2) scenario_inputs reference row ----------------------------------
    si = wb["scenario_inputs"]
    assert all(si.cell(APPEND_SCENARIO_ROW, c).value is None for c in (1, 2, 3))
    si.cell(APPEND_SCENARIO_ROW, 1, "CBO Feb-2026 (pub-61879, derived)")
    si.cell(APPEND_SCENARIO_ROW, 2, 573959)
    si.cell(APPEND_SCENARIO_ROW, 3, "≈ +50k/mo")

    # -- 3) _meta stamps ----------------------------------------------------
    meta = wb["_meta"]
    r = APPEND_META_FIRST_ROW
    assert meta.cell(r, 1).value is None
    for k, v, note in _META_KEYS:
        meta.cell(r, 1, k); meta.cell(r, 2, v); meta.cell(r, 3, note)
        r += 1

    # -- 4) control_vintage pointer -----------------------------------------
    cv = wb["control_vintage"]
    cv_row = cv.max_row + 2
    while any(cv.cell(cv_row, c).value is not None for c in range(1, 5)):
        cv_row += 1
    cv.cell(cv_row, 1, f"(re-port {REPORT_DATE}: gross_migration_sums tab added @ {PIN}; "
                       "population controls unchanged)")

    wb.save(TEMPLATE)

    # -- assertions on the saved file ---------------------------------------
    wb2 = openpyxl.load_workbook(TEMPLATE)
    assert len(wb2.sheetnames) == 10, wb2.sheetnames
    after = snapshot(wb2)
    allowed = {
        "scenario_inputs": {(APPEND_SCENARIO_ROW, c) for c in (1, 2, 3)},
        "_meta": {(rr, c) for rr in range(APPEND_META_FIRST_ROW, r)
                  for c in (1, 2, 3)},
        "control_vintage": {(cv_row, c) for c in (1,)},
    }
    for name in before:
        extra = set(after[name]) - set(before[name])
        assert extra <= allowed.get(name, set()), (name, sorted(extra - allowed.get(name, set()))[:5])
        for key, val in before[name].items():
            assert after[name].get(key) == val, (name, key, val, after[name].get(key))
    # yellow knob styling survives (the reader-consumed knob cell B4)
    assert wb2["scenario_inputs"].cell(4, 2).fill.start_color.rgb in ("FFFFFF00", "00FFFF00"), \
        wb2["scenario_inputs"].cell(4, 2).fill.start_color.rgb

    # sums-tab tie-outs (committed vintage)
    s = pd.read_excel(TEMPLATE, sheet_name="gross_migration_sums", header=3)
    d26 = s[s["year"] == 2026]
    i = d26.loc[d26.migration_flow == "immigration", "people"].sum()
    e = d26.loc[d26.migration_flow == "emigration", "people"].sum()
    i16 = d26.loc[(d26.migration_flow == "immigration") & (d26.bucket == "16plus"), "people"].sum()
    e16 = d26.loc[(d26.migration_flow == "emigration") & (d26.bucket == "16plus"), "people"].sum()
    assert (i, e) == (1_535_765, 961_806), (i, e)
    assert i - e == 573_959 and i16 - e16 == 499_825
    for got, want in ((i16 / i, 0.8940), (e16 / e, 0.9078), ((i16 - e16) / (i - e), 0.8708)):
        assert abs(got - want) <= 1e-4 + 1e-6, (got, want)
    print(f"template v2 OK: 10 tabs; {len(s)} sums rows; all appends in enumerated ranges; "
          "pre-existing cells identical; tie-outs reproduce")


if __name__ == "__main__":
    main(Path(sys.argv[1]))
