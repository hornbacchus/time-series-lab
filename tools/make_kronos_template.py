"""Generate the Kronos Forecast input template (.xlsx) -- run once; the output
is committed at engine/techniques/kronos_forecast/resources/templates/.

Data: the LINEAGE-VERIFIED example extract from the Kronos program
(data/examples/IEF_sample_250.csv beside KRONOS_TSL_HANDOFF.md): the last 250
trading days of the frozen v1 IEF snapshot. The sha256 is asserted here so a
regeneration can never silently pick up a different file -- byte-traceability
is checked, not assumed (extract c4b9342c..., source manifest 9bb16461...).

Layout contract (read by kronos_forecast/workbook_input.py):
  * the EXPERIMENTAL status block at the top (display only);
  * key/value parameter rows scanned from column A;
  * the data block headed by the row whose column-A cell is "Date".
"""

import hashlib
import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SRC = os.environ.get(
    "KRONOS_EXAMPLE_CSV",
    r"C:\Users\matth\OneDrive\Projects\Kronos Cross-Asset Forecaster"
    r"\data\examples\IEF_sample_250.csv")
EXPECTED_SHA = "c4b9342cdfafbf98c777697ade90caf42af85bab4f6a31383032d2142d34caef"
DST = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "engine", "techniques", "kronos_forecast", "resources",
                   "templates", "kronos_forecast_input_template.xlsx")

STATUS = ("EXPERIMENTAL -- research curiosity, not a validated forecaster. "
          "Out-of-sample evaluation (Jul 2025 - Jun 2026) found no confirmed "
          "predictive or distributional value; shaded bands are sampled path "
          "spread, not calibrated confidence intervals; model knowledge ends "
          "~June 2025 (it cannot know current events). Not for published "
          "research. Validation re-test 2026-12-01 (RATES_RERUN_PROTOCOL.md). "
          "Full record: KXAF_CLOSEOUT_MEMO.md.")

PARAMS = [
    ("lookback_L", 120, "context bars used (allowed 120-250)"),
    ("horizon_H", 10, "forecast trading days -- HARD CAP 25 (the tested envelope)"),
    ("paths_M", 10, "sampled paths -- cap 50. Latency ~= 11s load + M x H x 0.33s: "
                    "M=10,H=10 ~ 30-60s; M=10,H=25 ~ 1.5-2 min; M=50,H=25 ~ 5-7 min"),
    ("seed", 9000, "same inputs + seed -> identical outputs"),
    ("show_raw_ohlc", "FALSE", "TRUE adds a raw sampled-OHLC sheet with per-bar "
                               "validity flags (bars can be structurally invalid)"),
    ("T", 0.6, "DISPLAY-ONLY (fixed sampling temperature; edits ignored)"),
    ("top_p", 0.90, "DISPLAY-ONLY (fixed nucleus sampling; edits ignored)"),
]


def main():
    sha = hashlib.sha256(open(SRC, "rb").read()).hexdigest()
    if sha != EXPECTED_SHA:
        sys.exit(f"FATAL: {SRC} sha256 {sha[:12]}... != expected "
                 f"{EXPECTED_SHA[:12]}... -- refusing to build from "
                 f"unverified data.")
    df = pd.read_csv(SRC)

    wb = Workbook()
    ws = wb.active
    ws.title = "kronos_input"

    ws["A1"] = "Kronos Forecast -- EXPERIMENTAL input template"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = STATUS
    ws["A3"].font = Font(bold=True, color="9C0006")
    ws["A3"].fill = PatternFill("solid", fgColor="FFC7CE")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A3:F5")
    ws.row_dimensions[3].height = 70

    # The Bespoke input-cell convention (owner ruling, K3.2): every cell the
    # READER consumes gets solid yellow #FFFF00 + thin black borders on all
    # four sides. T/top_p are read-but-ignored display constants -> NOT styled
    # (the convention must not lie).
    input_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    thin = Side(style="thin", color="000000")
    input_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    r = 7
    ws.cell(row=r, column=1, value="Parameters (column B is yours to edit)").font = Font(bold=True)
    for name, val, note in PARAMS:
        r += 1
        ws.cell(row=r, column=1, value=name).font = Font(bold=True)
        v = ws.cell(row=r, column=2, value=val)
        if name not in ("T", "top_p"):
            v.fill = input_fill
            v.border = input_border
        ws.cell(row=r, column=3, value=note)

    r += 2
    hdr = r
    for ci, name in enumerate(["Date", "Open", "High", "Low", "Close", "Volume"], 1):
        c = ws.cell(row=hdr, column=ci, value=name)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    for i, row in df.iterrows():
        rr = hdr + 1 + i
        c0 = ws.cell(row=rr, column=1, value=pd.Timestamp(row["timestamps"]).date())
        c0.fill = input_fill
        c0.border = input_border
        for ci, col in enumerate(["open", "high", "low", "close", "volume"], 2):
            c = ws.cell(row=rr, column=ci, value=float(row[col]))
            c.fill = input_fill
            c.border = input_border
            # DISPLAY-ONLY number formats (K4 finding #2): prices at 3
            # decimals, volume as a grouped integer. Cell VALUES stay full
            # precision -- the lineage sha + the reproduction property depend
            # on values, never on display.
            c.number_format = "0.000" if col != "volume" else "#,##0"
    ws.cell(row=hdr + len(df) + 2, column=1,
            value="Any OHLCV series may be pasted over the block above "
                  "(>=120 rows; Volume may be blank). Example data: the last "
                  "250 trading days of IEF from the Kronos program's frozen, "
                  "sha-manifested snapshot (static; not a live feed).")

    # Column widths sized for the DATA content (K4 finding #1): the OHLC
    # price columns uniform; Date/Volume sensible for theirs. The parameter
    # annotations in C8:C14 deliberately do NOT drive any width -- they
    # OVERFLOW across the empty D..F cells of their rows (standard Excel
    # behavior; fully readable; no merged spans, so the reader's column-A/B
    # scan geometry is untouched).
    for col, w in (("A", 14), ("B", 11), ("C", 11), ("D", 11), ("E", 11), ("F", 12)):
        ws.column_dimensions[col].width = w

    os.makedirs(os.path.dirname(DST), exist_ok=True)
    wb.save(DST)
    print(f"written {DST} ({len(df)} data rows; source sha verified "
          f"{sha[:12]}...)")


if __name__ == "__main__":
    main()
